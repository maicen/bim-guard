"""Analysis routes for orchestrating compliance checks and rendering results."""

import os
from functools import lru_cache
from urllib.parse import quote

from fasthtml.common import (
    H3,
    A,
    Details,
    Div,
    FileResponse,
    Form,
    Iframe,
    Option,
    P,
    Request,
    Response,
    Script,
    Span,
    Summary,
    Table,
    Tbody,
    Td,
    Th,
    Thead,
    Title,
    Tr,
)
from monsterui.all import H1, Container
from starlette.responses import RedirectResponse, StreamingResponse

from app.components.layout import DashboardLayout
from app.components.ui import (
    Alert,
    AlertT,
    Card,
    CardContent,
    CardHeader,
    CardTitle,
    Checkbox,
    CountTableItemSpec,
    FormLabel,
    IconLinkButton,
    ItemsCountDataTable,
    Label,
    Select,
    SubmitButton,
)
from app.components.ui import (
    Table as UiTable,
)
from app.components.ui import (
    TableBody as UiTableBody,
)
from app.components.ui import (
    TableCell as UiTableCell,
)
from app.components.ui import (
    TableHead as UiTableHead,
)
from app.components.ui import (
    TableHeader as UiTableHeader,
)
from app.components.ui import (
    TableRow as UiTableRow,
)
from app.logging_config import get_logger
from app.modules.orchestrator import BIMGuard_App
from app.services.documents_service import DocumentService
from app.services.projects_service import ProjectsService
from app.services.report_artifacts import ReportArtifactService
from app.services.rules_service import RuleService

_bim_guard_app = BIMGuard_App()
_projects_service = ProjectsService()
_documents_service = DocumentService()
_rule_service = RuleService()
_report_artifact_service = ReportArtifactService()
logger = get_logger(__name__)

_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "data")

# Rule-compliance results from the last Initial Analysis run per project, so
# the IFC graph iframe can reuse them for violation highlighting instead of
# re-running the full IFC-parse + compliance pipeline (which alone takes
# 100s+ on a real building model — measured on bimguard_headquarter1).
_last_initial_rule_compliance: dict[int, list[dict]] = {}

# Project ID for the ARCH-analysis card render currently in flight — read by
# _simple_summary_table()'s "View in 3D" links. Threading project_id as an
# explicit parameter would mean touching all 9 _simple_*_card() wrappers plus
# _simple_domain_card() just to pass one value through; this follows the same
# last-result-cache convention already used above and by _last_compliance_results.
_last_summary_project_id: int | None = None


@lru_cache(maxsize=1)
def _code_rule_context() -> dict:
    """Return cached code-reference labels/limits resolved from saved code rules."""
    context = {
        "daylight_ref": "applicable code rule",
        "daylight_ratio_label": "1/10",
        "fire_ref": "applicable code rule",
        "fire_min_label": "45 min",
        "garage_ref": "applicable code rule",
        "exit_ref": "applicable code rule",
        "travel_ref": "applicable code rule",
        "travel_max_label": "25 m",
        "refs_by_target": {},
    }

    try:
        rules = _rule_service.list_code_rules()
    except Exception:
        return context

    def _as_float(value):
        if isinstance(value, (int, float)):
            return float(value)
        if value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    refs_by_target: dict[str, list[str]] = {}
    for row in rules:
        target = str(row.get("target_ifc_class") or "").strip()
        ref = str(row.get("reference") or "").strip()
        if not target or not ref:
            continue
        bucket = refs_by_target.setdefault(target, [])
        if ref not in bucket:
            bucket.append(ref)
    context["refs_by_target"] = refs_by_target

    for row in rules:
        ref = str(row.get("reference") or "")
        if "9.7.2" not in ref:
            continue
        context["daylight_ref"] = ref
        ratio = _as_float(row.get("check_value"))
        unit = str(row.get("unit") or "").strip().lower()
        if ratio and unit == "ratio" and ratio > 0:
            context["daylight_ratio_label"] = f"1/{int(round(1 / ratio))}"
            break

    for row in rules:
        ref = str(row.get("reference") or "")
        if "9.10.9" not in ref:
            continue
        if str(row.get("target_ifc_class") or "") != "IfcWall":
            continue
        if str(row.get("operator") or "") != ">=":
            continue
        if str(row.get("property_name") or "") not in (
            "FireRating",
            "FireResistanceRating",
            "FireResistance",
            "REI",
            "FRR",
        ):
            continue
        min_v = _as_float(row.get("check_value"))
        if min_v is None:
            continue
        context["fire_ref"] = ref
        min_label = str(int(min_v)) if float(min_v).is_integer() else str(min_v)
        unit = str(row.get("unit") or "min").strip() or "min"
        context["fire_min_label"] = f"{min_label} {unit}"
        break

    for row in rules:
        ref = str(row.get("reference") or "")
        if "9.10.14.2" in ref:
            context["garage_ref"] = ref
            break

    for row in rules:
        ref = str(row.get("reference") or "")
        if "9.9.4.1" in ref:
            context["exit_ref"] = ref
            break

    for row in rules:
        ref = str(row.get("reference") or "")
        if "9.9.10.1" not in ref:
            continue
        context["travel_ref"] = ref
        max_v = _as_float(row.get("check_value"))
        if max_v is not None:
            max_label = str(int(max_v)) if float(max_v).is_integer() else str(max_v)
            unit = str(row.get("unit") or "m").strip() or "m"
            context["travel_max_label"] = f"{max_label} {unit}"
        break

    return context


def _section_title(label: str, ref: str) -> str:
    """Return a standard section title suffixing a resolved code reference."""
    return f"{label} - {ref}" if ref else label


def _result_ref(result: dict) -> str:
    """Read a normalized rule reference from module outputs."""
    return str(result.get("code_ref") or "").strip()


def _domain_ref_for_targets(targets: list[str], fallback: str = "") -> str:
    """Return a combined code-reference label for one or more IFC targets."""
    refs_by_target = _code_rule_context().get("refs_by_target") or {}
    refs: list[str] = []
    for target in targets:
        for ref in refs_by_target.get(target, []):
            if ref not in refs:
                refs.append(ref)
    return " · ".join(refs) if refs else fallback


def _active_code_refs_summary() -> str:
    """Return a short, user-facing summary of active code references."""
    ctx = _code_rule_context()
    refs: list[str] = []
    for ref in (
        ctx.get("daylight_ref"),
        ctx.get("fire_ref"),
        ctx.get("garage_ref"),
        ctx.get("exit_ref"),
        ctx.get("travel_ref"),
    ):
        ref_s = str(ref or "").strip()
        if not ref_s or ref_s == "applicable code rule" or ref_s in refs:
            continue
        refs.append(ref_s)
    if not refs:
        return "saved code references"
    return " · ".join(refs[:3]) + (" · ..." if len(refs) > 3 else "")


def _analysis_form(projects, documents, mode: str):
    """Build the analysis form for the requested workflow mode."""
    is_simple = mode == "simple"
    is_initial = mode == "initial"
    project_options = [Option("— select a project —", value="", disabled=True, selected=True)] + [
        Option(p.get("name", f"Project {p['id']}"), value=str(p["id"])) for p in projects
    ]

    form_sections = [
        Div(
            FormLabel("Project (IFC Model)", fr="project_id"),
            Select(
                *project_options,
                id="project_id",
                name="project_id",
                required=True,
            ),
        ),
    ]

    # Rule Folder — shown for every mode, right under the project picker, so
    # you can scope a check to rules you extracted yourself instead of the
    # built-in default code rules seeded from source (code_seed_rules.py).
    folders = _rule_service.list_folders()
    folder_options = [Option("All folders", value="", selected=True)] + [
        Option(f"{f['ruleset_id']} ({f['count']})", value=f["ruleset_id"]) for f in folders
    ]
    if is_simple:
        folder_help = (
            "Narrow the check to one saved folder of rules — only that folder's rules "
            "are used, the built-in seeded default code rules are excluded. Leave "
            "on 'All folders' to test everything, including the built-in defaults."
        )
    elif mode == "model-rules":
        folder_help = (
            "Narrow the check to one saved MEP ruleset folder (for example GC-001, "
            "CC-001, or MC-001). Leave on 'All folders' to run against all MEP "
            "rules available in the library."
        )
    else:
        folder_help = (
            "Narrow the check to one saved folder of rules — only that folder's "
            "rules are used, the built-in seeded default code rules are excluded. Leave on "
            "'All folders' to test every rule in the theme above, including "
            "built-in defaults."
        )
    form_sections.append(
        Div(
            FormLabel("Rule Folder", fr="rule_folder"),
            Select(
                *folder_options,
                id="rule_folder",
                name="rule_folder",
            ),
            P(folder_help, cls="text-xs text-muted-foreground mt-1"),
        )
    )

    if mode == "initial":
        form_sections.append(
            Div(
                FormLabel("Analysis Theme", fr="analysis_theme"),
                Select(
                    Option("Architecture", value="Architecture", selected=True),
                    Option("MEP", value="MEP"),
                    id="analysis_theme",
                    name="analysis_theme",
                    required=True,
                ),
            )
        )
    elif mode == "model-rules":
        form_sections.append(
            Div(
                FormLabel("MEP Engines"),
                Div(
                    Span(
                        "GC-001 Galvanic",
                        cls="inline-block px-2 py-0.5 rounded text-xs font-semibold bg-blue-100 text-blue-800",
                    ),
                    Span(
                        "CC-001 Crevice",
                        cls="inline-block px-2 py-0.5 rounded text-xs font-semibold bg-teal-100 text-teal-800",
                    ),
                    Span(
                        "MC-001 MIC",
                        cls="inline-block px-2 py-0.5 rounded text-xs font-semibold bg-indigo-100 text-indigo-800",
                    ),
                    cls="flex flex-wrap items-center gap-2",
                ),
                P(
                    "All three MEP corrosion engines are included in this workflow.",
                    cls="text-xs text-muted-foreground mt-1",
                ),
            )
        )

    if not is_initial and not is_simple:
        doc_checkboxes = []
        for doc in documents:
            doc_checkboxes.append(
                Div(
                    Checkbox(
                        id=f"doc_{doc['id']}",
                        name="document_ids",
                        value=str(doc["id"]),
                        cls="mr-2",
                    ),
                    Label(
                        doc.get("filename", f"Document {doc['id']}"),
                        for_=f"doc_{doc['id']}",
                        cls="text-sm cursor-pointer",
                    ),
                    cls="flex items-center gap-1",
                )
            )

        if not doc_checkboxes:
            doc_checkboxes = [P("No documents uploaded yet.", cls="text-sm text-muted-foreground")]

        form_sections.append(
            Div(
                FormLabel("Documents"),
                Div(
                    *doc_checkboxes,
                    cls="space-y-2 border rounded-md p-3 bg-muted/30",
                ),
            )
        )

    if not is_simple:
        form_sections.append(
            Div(
                FormLabel("Count Options"),
                Div(
                    Div(
                        Checkbox(
                            id="include_openings",
                            name="include_openings",
                            value="1",
                            checked=True,
                            cls="mr-2",
                        ),
                        Label(
                            "Include openings (IfcOpeningElement)",
                            for_="include_openings",
                            cls="text-sm cursor-pointer",
                        ),
                        cls="flex items-center gap-1",
                    ),
                    Div(
                        Checkbox(
                            id="include_spaces",
                            name="include_spaces",
                            value="1",
                            checked=True,
                            cls="mr-2",
                        ),
                        Label(
                            "Include spaces (IfcSpace)",
                            for_="include_spaces",
                            cls="text-sm cursor-pointer",
                        ),
                        cls="flex items-center gap-1",
                    ),
                    Div(
                        Checkbox(
                            id="include_type_definitions",
                            name="include_type_definitions",
                            value="1",
                            cls="mr-2",
                        ),
                        Label(
                            "Include type definitions (IfcElementType)",
                            for_="include_type_definitions",
                            cls="text-sm cursor-pointer",
                        ),
                        cls="flex items-center gap-1",
                    ),
                    cls="space-y-2 border rounded-md p-3 bg-muted/30",
                ),
            )
        )

    if mode == "simple":
        submit_label = "Run ARCH Analysis"
        results_id = "simple-analysis-results"
        hx_post = "/analysis/ARCH/results"
        card_title = "ARCH"
        helper_copy = (
            "Check the IFC model against saved code categories: "
            "Windows, Doors, Stairs, Egress, Fire Protection, Garage, and more. "
            f"Active references: {_active_code_refs_summary()}."
        )
    elif is_initial:
        submit_label = "Run Initial Analysis"
        results_id = "initial-analysis-results"
        hx_post = "/analysis/initial/results"
        card_title = "Initial Analysis"
        helper_copy = (
            "Inspect the IFC model structure, counts, and quality signals before checking rules."
        )
    else:
        submit_label = "Run MEP Analysis"
        results_id = "model-rules-results"
        hx_post = "/analysis/results"
        card_title = "MEP"
        helper_copy = (
            "Run MEP corrosion-focused analysis using the saved MEP rules in the "
            "library (GC-001, CC-001, and MC-001)."
        )

    loader_id = f"{results_id}-loader"

    # JS: on submit → copy loading card into results area + disable button
    before_js = (
        f"var r=document.getElementById('{results_id}'),"
        f"l=document.getElementById('{loader_id}');"
        "if(r&&l)r.innerHTML=l.innerHTML;"
        "var b=this.querySelector('[type=submit]');"
        "if(b){{b.disabled=true;b.dataset.orig=b.textContent;b.textContent='Analysing…';}}"
    )
    after_js = (
        "var b=this.querySelector('[type=submit]');"
        "if(b){{b.disabled=false;if(b.dataset.orig)b.textContent=b.dataset.orig;}}"
    )

    return Div(
        Card(
            CardHeader(
                Div(
                    CardTitle(card_title),
                    P(helper_copy, cls="text-sm text-muted-foreground mt-1"),
                )
            ),
            CardContent(
                Form(
                    Div(
                        *form_sections,
                        Div(
                            SubmitButton(submit_label, variant="primary"),
                            cls="flex items-center gap-4",
                        ),
                    ),
                    method="post",
                    action=hx_post,
                    hx_post=hx_post,
                    hx_target=f"#{results_id}",
                    hx_swap="innerHTML",
                    **{
                        "hx-on:htmx:before-request": before_js,
                        "hx-on:htmx:after-request": after_js,
                    },
                ),
            ),
        ),
        # Hidden loading card — cloned into results area when request starts
        Div(
            Div(
                Div(
                    style=(
                        "width:40px;height:40px;border-radius:50%;"
                        "border:4px solid #e2e8f0;border-top-color:#3b82f6;"
                        "animation:bimguard-spin .75s linear infinite;"
                        "margin:0 auto 16px;"
                    )
                ),
                P("Analysing model…", cls="text-base font-semibold"),
                P(
                    "Loading IFC · Extracting properties · Running compliance checks",
                    cls="text-xs text-muted-foreground mt-1",
                ),
                cls="text-center py-14",
            ),
            cls="border rounded-lg shadow-sm bg-card",
            id=loader_id,
            style="display:none",
        ),
        Script(
            "(function(){var s=document.createElement('style');"
            "s.textContent='@keyframes bimguard-spin{to{transform:rotate(360deg)}}';"
            "document.head.appendChild(s);})();"
        ),
        Div(id=results_id),
    )


async def _run_analysis_request(req: Request, forced_theme: str | None = None):
    """Parse and validate the analysis request, then run the orchestrator."""
    form = await req.form()
    project_id_raw = form.get("project_id") or ""
    if not project_id_raw:
        logger.warning("Rejected analysis request without a project ID")
        return None, Alert("Please select a project.", cls=AlertT.error)
    try:
        project_id = int(project_id_raw)
    except ValueError:
        logger.warning("Rejected analysis request with invalid project ID=%r", project_id_raw)
        return None, Alert("Invalid project selection.", cls=AlertT.error)

    doc_ids = [int(v) for v in form.getlist("document_ids") if v]
    analysis_theme = (forced_theme or form.get("analysis_theme") or "Architecture").strip()
    rule_folder = (form.get("rule_folder") or "").strip()
    include_openings = bool(form.get("include_openings"))
    include_spaces = bool(form.get("include_spaces"))
    include_type_definitions = bool(form.get("include_type_definitions"))
    logger.info(
        "Starting analysis request project_id=%d theme=%s documents=%d rule_folder=%s",
        project_id,
        analysis_theme,
        len(doc_ids),
        rule_folder or "all",
    )
    result = _bim_guard_app.orchestrate_workflow(
        project_id,
        doc_ids,
        analysis_theme=analysis_theme,
        rule_folder=rule_folder,
        include_openings=include_openings,
        include_spaces=include_spaces,
        include_type_definitions=include_type_definitions,
    )

    if "error" in result:
        logger.warning("Analysis request failed project_id=%d error=%s", project_id, result["error"])
        return None, Alert(result["error"], cls=AlertT.error)

    try:
        _report_artifact_service.persist_bcf(project_id, result.get("bcf_topics", []))
    except Exception:
        logger.exception("BCF persistence failed project_id=%d", project_id)
        return None, Alert(
            "Analysis completed, but the BCF report could not be saved. Please try again.",
            cls=AlertT.error,
        )

    logger.info(
        "Analysis request complete project_id=%d ifc_elements=%d rule_results=%d",
        project_id,
        result.get("ifc_element_count", 0),
        len(result.get("rule_compliance", [])),
    )
    return {"project_id": project_id, "result": result}, None


def _build_ifc_summary_content(result: dict, project_id: int):
    """Build the IFC detail block shared by the initial analysis page."""
    ifc_count = result["ifc_element_count"]
    ifc_type_counts = result.get("ifc_type_counts") or {}
    ifc_totals = result.get("ifc_totals") or {}
    ifc_error = result["ifc_error"]
    ifc_quality = result.get("ifc_quality_report") or {}
    ifc_quality_warnings = result.get("ifc_quality_warnings") or []
    ifc_quality_improvements = result.get("ifc_quality_improvements") or []
    ifc_schema_note = result.get("ifc_schema_note")

    if ifc_error:
        return P(f"IFC parsing error: {ifc_error}", cls="text-sm text-destructive")
    if not _projects_service.resolve_ifc_file(project_id):
        return P(
            "No IFC file attached to this project.",
            cls="text-sm text-muted-foreground",
        )

    filters = ifc_totals.get("filters") or {}
    deltas = ifc_totals.get("excluded_or_added") or {}
    overall = ifc_quality.get("overall") or {}
    labeling = ifc_quality.get("labeling") or {}
    guids = ifc_quality.get("guids") or {}
    properties = ifc_quality.get("properties") or {}

    quality_alerts = [
        Alert(msg, cls=AlertT.warning if hasattr(AlertT, "warning") else "")
        for msg in ifc_quality_warnings
    ]
    if ifc_schema_note:
        quality_alerts.append(
            Alert(
                ifc_schema_note,
                cls=AlertT.info if hasattr(AlertT, "info") else "",
            )
        )

    quality_block = (
        _collapsible_card(
            "IFC Quality",
            P(f"Overall score: {overall.get('score', 0):.1f}%", cls="text-sm font-medium"),
            Div(
                P(
                    f"Labeling: {labeling.get('score', 0):.1f}%",
                    cls="text-xs text-muted-foreground",
                ),
                P(f"GUIDs: {guids.get('score', 0):.1f}%", cls="text-xs text-muted-foreground"),
                P(
                    f"Properties: {properties.get('score', 0):.1f}%",
                    cls="text-xs text-muted-foreground",
                ),
                cls="space-y-1 mt-2",
            ),
            *quality_alerts,
            open=False,
        )
        if ifc_quality
        else ""
    )

    improvement_block = (
        _collapsible_card(
            "Improvement Summary",
            Div(
                *[
                    P(message, cls="text-sm text-muted-foreground")
                    for message in ifc_quality_improvements
                ],
                cls="space-y-1",
            ),
            open=False,
        )
        if ifc_quality_improvements
        else ""
    )

    counts_table = ItemsCountDataTable(
        [
            CountTableItemSpec(
                label="Built Elements",
                total=int(ifc_totals.get("built_elements", ifc_count)),
                subtotal=int(ifc_totals.get("built_elements", ifc_count)),
                note="Schema-aware building entities",
            ),
            CountTableItemSpec(
                label="All Physical Elements",
                total=int(ifc_totals.get("all_physical_elements", 0)),
                subtotal=int(ifc_totals.get("adjusted_physical_elements", ifc_count)),
                note="Based on IfcElement",
            ),
            CountTableItemSpec(
                label="All Products",
                total=int(ifc_totals.get("all_products", 0)),
                subtotal=int(ifc_totals.get("adjusted_products", 0)),
                note="Based on IfcProduct",
            ),
        ],
        caption="Built, physical, and product item counts",
        options_summary=(
            "Options: "
            f"openings={'on' if filters.get('include_openings', True) else 'off'} "
            f"(count: {deltas.get('openings', 0)}), "
            f"spaces={'on' if filters.get('include_spaces', True) else 'off'} "
            f"(count: {deltas.get('spaces', 0)}), "
            f"type defs={'on' if filters.get('include_type_definitions', False) else 'off'} "
            f"(count: {deltas.get('type_definitions', 0)})."
        ),
        built_type_breakdown=ifc_type_counts,
    )

    return Div(
        quality_block,
        improvement_block,
        counts_table,
        cls="space-y-1",
    )


def _build_document_cards(result: dict):
    """Build document summary cards for the model-vs-rules page."""
    cards = []
    for doc in result["documents"]:
        cards.append(
            Card(
                CardHeader(CardTitle(doc["filename"] or "Untitled document")),
                CardContent(
                    P(
                        f"{doc['section_count']} text sections extracted.",
                        cls="text-sm text-muted-foreground",
                    )
                ),
            )
        )
    return cards


def _build_ifc_graph_card(result: dict, project_id: int):
    """Embed the PyVis IFC relationship graph (containment/aggregation/connectivity)."""
    if (
        not project_id
        or result.get("ifc_error")
        or not _projects_service.resolve_ifc_file(project_id)
    ):
        return Card(
            CardHeader(CardTitle("IFC Relationship Graph")),
            CardContent(
                P(
                    "No IFC file available for this project — upload one to see the "
                    "relationship graph.",
                    cls="text-sm text-muted-foreground",
                )
            ),
        )

    src = f"/reports/ifc-graph/{project_id}"

    return Card(
        CardHeader(CardTitle("IFC Relationship Graph")),
        CardContent(
            P(
                "Containment, aggregation, and connectivity relationships between "
                "elements. Red nodes failed a compliance rule; green nodes are "
                "spatial containers (project/site/building/storey/space).",
                cls="text-xs text-muted-foreground mb-2",
            ),
            Iframe(
                src=src,
                style=("width:100%;height:720px;border:1px solid var(--border);border-radius:8px;"),
            ),
        ),
    )


def _band_badge(band: str):
    colours = {
        "Critical": "bg-red-600 text-white",
        "High": "bg-orange-500 text-white",
        "Medium": "bg-yellow-400 text-black",
        "Low": "bg-green-600 text-white",
    }
    return Span(
        band,
        cls=f"inline-block px-2 py-0.5 rounded text-xs font-semibold {colours.get(band, 'bg-gray-400 text-white')}",
    )


def _compliance_card(results, cost_impact, issue_stats, is_demo, project_id, error):
    """Build the corrosion compliance results card for the analysis results page."""
    if error:
        return Card(
            CardHeader(CardTitle("Corrosion Compliance — GC-001 / CC-001 / MC-001")),
            CardContent(P(f"Compliance engine error: {error}", cls="text-sm text-destructive")),
        )

    if not results:
        return None

    bands = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    for r in results:
        b = r.get("risk_band", "Low")
        if b in bands:
            bands[b] += 1

    badge_row = Div(
        *[
            Div(
                _band_badge(b),
                Span(f" {bands[b]}", cls="text-sm font-medium ml-1"),
                cls="flex items-center",
            )
            for b in ("Critical", "High", "Medium", "Low")
        ],
        cls="flex items-center gap-4 flex-wrap",
    )

    cost_line = (
        P(
            f"Estimated remediation: £{cost_impact.total_cost_gbp:,.0f}  |  "
            f"Programme impact: {cost_impact.total_days} days",
            cls="text-sm text-muted-foreground mt-2",
        )
        if cost_impact
        else ""
    )

    tracker_line = (
        P(
            f"Issue history: {issue_stats.get('new', 0)} new, "
            f"{issue_stats.get('updated', 0)} updated, "
            f"{issue_stats.get('resolved', 0)} resolved.",
            cls="text-xs text-muted-foreground mt-1",
        )
        if issue_stats
        else ""
    )

    demo_notice = (
        Alert(
            "No IFC file found — showing synthetic demo data (25 representative MEP elements).",
            cls=AlertT.info if hasattr(AlertT, "info") else "",
        )
        if is_demo
        else ""
    )

    flagged = [r for r in results if r.get("risk_band", "Low") != "Low"]

    if flagged:
        header_cells = [
            Th(h, cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted")
            for h in (
                "Element",
                "Floor",
                "Material",
                "Engine",
                "Band",
                "Score",
                "Required Action",
            )
        ]
        data_rows = []
        for r in flagged[:20]:
            data_rows.append(
                Tr(
                    Td(r.get("name", "—")[:40], cls="px-3 py-2 text-sm"),
                    Td(r.get("floor", "—"), cls="px-3 py-2 text-sm"),
                    Td(r.get("material_a", "—")[:22], cls="px-3 py-2 text-sm"),
                    Td(
                        Span(
                            str(r.get("dominant_mechanism", "—")).upper(),
                            cls="inline-block px-2 py-0.5 rounded text-xs font-semibold bg-slate-100 text-slate-700",
                        ),
                        cls="px-3 py-2",
                    ),
                    Td(_band_badge(r.get("risk_band", "Low")), cls="px-3 py-2"),
                    Td(f"{r.get('overall_score', 0):.3f}", cls="px-3 py-2 text-sm font-mono"),
                    Td(r.get("action", "—")[:70], cls="px-3 py-2 text-xs"),
                    cls="border-b border-muted last:border-0",
                )
            )
        if len(flagged) > 20:
            data_rows.append(
                Tr(
                    Td(
                        f"… and {len(flagged) - 20} more flagged elements",
                        cls="px-3 py-2 text-xs text-muted-foreground italic",
                        colspan="7",
                    )
                )
            )
        results_table = Div(
            Table(
                Thead(Tr(*header_cells)),
                Tbody(*data_rows),
                cls="w-full text-sm",
            ),
            cls="overflow-auto mt-4 border rounded-md",
        )
    else:
        results_table = P(
            "No elements flagged at Medium risk or above.",
            cls="text-sm text-muted-foreground mt-3",
        )

    bcf_btn = (
        Div(
            A(
                "Download BCF Report",
                href=f"/reports/bcf/{project_id}",
                cls="inline-block mt-4 px-4 py-2 bg-primary text-primary-foreground text-sm rounded-md hover:opacity-90",
            ),
        )
        if project_id
        else ""
    )

    return _collapsible_card(
        "Corrosion Compliance — GC-001 / CC-001 / MC-001",
        demo_notice,
        badge_row,
        cost_line,
        tracker_line,
        results_table,
        bcf_btn,
    )


def _mep_engine_rules_card():
    """Show MEP engine and ruleset coverage for the MEP analysis page."""
    engines = [
        ("GC-001", "Galvanic Corrosion", "BIMGUARD-GC-001", "bg-blue-100 text-blue-800"),
        ("CC-001", "Crevice Corrosion", "BIMGUARD-CC-001", "bg-teal-100 text-teal-800"),
        (
            "MC-001",
            "Microbially Influenced Corrosion",
            "BIMGUARD-MC-001",
            "bg-indigo-100 text-indigo-800",
        ),
    ]

    rows = []
    for mech, label, ruleset_id, badge_cls in engines:
        count = len(_rule_service.list_by_ruleset(ruleset_id))
        rows.append(
            Tr(
                Td(
                    Span(
                        mech,
                        cls=f"inline-block px-2 py-0.5 rounded text-xs font-semibold {badge_cls}",
                    ),
                    cls="px-3 py-2",
                ),
                Td(label, cls="px-3 py-2 text-sm"),
                Td(ruleset_id, cls="px-3 py-2 text-xs font-mono"),
                Td(str(count), cls="px-3 py-2 text-sm font-semibold"),
                cls="border-b border-muted last:border-0",
            )
        )

    return Card(
        CardHeader(CardTitle("MEP Engines and Rule Coverage")),
        CardContent(
            P(
                "MEP analysis runs corrosion, crevice, and MIC engines using their saved rule libraries.",
                cls="text-sm text-muted-foreground mb-3",
            ),
            Div(
                Table(
                    Thead(
                        Tr(
                            Th(
                                "Engine",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Scope",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Ruleset",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Rules",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                        )
                    ),
                    Tbody(*rows),
                    cls="w-full text-sm",
                ),
                cls="overflow-auto border rounded-md",
            ),
        ),
    )


def _rule_validation_card(rule_validations: list[dict], analysis_theme: str):
    """Build the Module 3 rule validation card for the analysis results page."""
    if not rule_validations:
        return Card(
            CardHeader(CardTitle(f"Rule Validation — Module 3 ({analysis_theme})")),
            CardContent(
                P(
                    f"No matching rules found in the library for {analysis_theme}. "
                    "Go to Library → Rule Extraction Studio to extract and save rules first.",
                    cls="text-sm text-muted-foreground",
                )
            ),
        )

    present = sum(1 for r in rule_validations if r["status"] == "present")
    not_found = len(rule_validations) - present

    summary = Div(
        Div(
            Span(str(len(rule_validations)), cls="text-2xl font-bold"),
            Span(" rules checked", cls="text-sm text-muted-foreground ml-1"),
            cls="flex items-baseline",
        ),
        Div(
            Span(
                f"{present} matched",
                cls="inline-block px-2 py-0.5 rounded text-xs font-semibold bg-green-100 text-green-800 mr-2",
            ),
            Span(
                f"{not_found} no elements",
                cls="inline-block px-2 py-0.5 rounded text-xs font-semibold bg-yellow-100 text-yellow-800",
            ),
            cls="mt-1",
        ),
        cls="mb-4",
    )

    header_cells = [
        Th(h, cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted")
        for h in ("Reference", "Description", "Target IFC Class", "Elements in Model", "Status")
    ]

    def _status_badge(status: str, count: int):
        if status == "present":
            return Span(
                f"✓ {count} found",
                cls="inline-block px-2 py-0.5 rounded text-xs font-semibold bg-green-100 text-green-800",
            )
        return Span(
            "No elements",
            cls="inline-block px-2 py-0.5 rounded text-xs font-semibold bg-yellow-100 text-yellow-800",
        )

    data_rows = []
    for r in rule_validations:
        data_rows.append(
            Tr(
                Td(r.get("reference", "—"), cls="px-3 py-2 text-xs font-mono"),
                Td(
                    r.get("description", "")[:80]
                    + ("…" if len(r.get("description", "")) > 80 else ""),
                    cls="px-3 py-2 text-sm",
                ),
                Td(r.get("target_ifc_class", "—"), cls="px-3 py-2 text-xs font-mono text-blue-700"),
                Td(str(r.get("element_count", 0)), cls="px-3 py-2 text-sm text-center"),
                Td(_status_badge(r["status"], r["element_count"]), cls="px-3 py-2"),
                cls="border-b border-muted last:border-0",
            )
        )

    return _collapsible_card(
        f"Rule Validation — Module 3 ({analysis_theme})",
        summary,
        Div(
            Table(
                Thead(Tr(*header_cells)),
                Tbody(*data_rows),
                cls="w-full text-sm",
            ),
            cls="overflow-auto border rounded-md",
        ),
        open=False,
    )


def _rule_compliance_card(
    compliance_results: list[dict],
    summary: dict,
    error: str | None,
    analysis_theme: str,
    project_id: int | None = None,
):
    """Module 4 rule compliance card — shows PASS/FAIL per rule with element details."""
    title = f"Rule Compliance Check — Module 4 ({analysis_theme})"

    if error:
        return Card(
            CardHeader(CardTitle(title)),
            CardContent(P(f"Compliance check error: {error}", cls="text-sm text-destructive")),
        )

    if not compliance_results:
        return None

    # Looked up once per render (not per failing element) — every "View in
    # 3D" link on this card points at the same run's freshly-persisted BCF
    # artifact, which _run_analysis_request() creates right before this
    # card is built.
    bcf_artifact_id = None
    if project_id is not None:
        latest = _report_artifact_service.latest_bcf(project_id)
        bcf_artifact_id = latest.get("id") if latest else None

    passed = summary.get("passed", 0)
    failed = summary.get("failed", 0)
    missing = summary.get("missing_data", 0)
    no_elem = summary.get("no_elements", 0)
    mand_f = summary.get("mandatory_failed", 0)

    summary_bar = Div(
        _pass_rate_pill(summary),
        Span(f"✓ {passed} passed", cls="text-xs text-green-700 mr-2"),
        Span(f"✗ {failed} failed", cls="text-xs text-red-700 mr-2"),
        Span(f"~ {missing} missing data", cls="text-xs text-yellow-700 mr-2") if missing else "",
        Span(f"○ {no_elem} no elements", cls="text-xs text-muted-foreground") if no_elem else "",
        Span(f"  ⚠ {mand_f} mandatory failures", cls="text-xs text-red-600 font-semibold ml-2")
        if mand_f
        else "",
        cls="flex flex-wrap items-center gap-1 mb-4",
    )

    _STATUS_CLS = {
        "PASS": "bg-green-100 text-green-800",
        "FAIL": "bg-red-100 text-red-800",
        "MISSING_DATA": "bg-yellow-100 text-yellow-800",
        "PARTIAL": "bg-orange-100 text-orange-800",
        "NO_ELEMENTS": "bg-gray-100 text-gray-600",
    }

    _IFC_LABELS = {
        "IfcStairFlight": "Stairs",
        "IfcDoor": "Doors",
        "IfcWindow": "Windows",
        "IfcRailing": "Railings & Guards",
        "IfcRamp": "Ramps",
        "IfcSlab": "Slabs & Landings",
        "IfcWall": "Walls",
        "IfcSpace": "Spaces & Rooms",
        "IfcZone": "Zones",
        "IfcColumn": "Columns",
        "IfcBeam": "Beams",
        "IfcFooting": "Footings & Foundations",
        "IfcPipeSegment": "Pipes",
        "IfcDuctSegment": "Ducts",
        "IfcSanitaryTerminal": "Plumbing Fixtures",
        "IfcAlarm": "Alarms & Detectors",
    }

    # Group results by IFC class, preserving order of first appearance
    from collections import defaultdict

    grouped: dict[str, list[dict]] = defaultdict(list)
    seen_targets: list[str] = []
    for r in compliance_results:
        t = r.get("target", "Other")
        if t not in grouped:
            seen_targets.append(t)
        grouped[t].append(r)

    header_cells = [
        Th(h, cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted")
        for h in ("Ref", "Description", "Property", "Rule", "Status", "Pass / Fail / Miss")
    ]

    rows = []
    for target in seen_targets:
        label = _IFC_LABELS.get(target, target)
        group_results = grouped[target]

        # Count group-level status summary
        g_pass = sum(1 for r in group_results if r.get("status") == "PASS")
        g_fail = sum(1 for r in group_results if r.get("status") == "FAIL")
        g_miss = sum(1 for r in group_results if r.get("status") in ("MISSING_DATA", "PARTIAL"))
        g_noelem = sum(1 for r in group_results if r.get("status") == "NO_ELEMENTS")

        badge_cls = (
            "bg-red-100 text-red-700"
            if g_fail
            else "bg-yellow-100 text-yellow-700"
            if g_miss
            else "bg-gray-100 text-gray-500"
            if g_noelem
            else "bg-green-100 text-green-700"
        )
        badge_txt = (
            f"{g_fail} failed"
            if g_fail
            else f"{g_miss} missing"
            if g_miss
            else "no elements"
            if g_noelem
            else "all pass"
        )

        rows.append(
            Tr(
                Td(
                    Div(
                        Span(label, cls="font-semibold text-sm text-foreground"),
                        Span(target, cls="text-xs text-muted-foreground font-mono ml-2"),
                        Span(
                            badge_txt,
                            cls=f"ml-auto text-xs px-2 py-0.5 rounded-full font-medium {badge_cls}",
                        ),
                        cls="flex items-center gap-2",
                    ),
                    colspan="6",
                    cls="px-3 py-2 bg-muted/60 border-t-2 border-border",
                ),
                cls="",
            )
        )

        for r in group_results:
            op = r.get("operator", "")
            if op == "between":
                rule_str = f"between {r.get('value_min')}–{r.get('value_max')} {r.get('unit', '')}"
            elif op == "exists":
                rule_str = "exists"
            else:
                rule_str = f"{op} {r.get('check_value', '')} {r.get('unit', '')}".strip()

            status = r.get("status", "")
            s_cls = _STATUS_CLS.get(status, "bg-gray-100 text-gray-600")

            counts = Span(
                f"{r.get('pass_count', 0)} / {r.get('fail_count', 0)} / {r.get('missing_count', 0)}",
                cls="font-mono text-xs",
            )

            # Collapsible failure details
            failures = r.get("failures", [])
            fail_detail = ""
            if failures:

                def _fmt_actual(v):
                    if isinstance(v, float):
                        return f"{v:,.2f}" if v >= 1 else f"{v:.4f}"
                    return str(v) if v is not None else "—"

                def _view_in_3d_link(f: dict):
                    guid = f.get("guid")
                    if project_id is None or bcf_artifact_id is None or not guid:
                        return ""
                    href = (
                        f"/viewer?project_id={project_id}&bcf_artifact_id={bcf_artifact_id}"
                        f"&element_guid={quote(guid)}"
                    )
                    return A(
                        "View in 3D",
                        href=href,
                        target="_blank",
                        cls="text-xs text-blue-700 hover:underline whitespace-nowrap",
                    )

                fail_rows = [
                    Tr(
                        Td(f.get("element_name", "—")[:35], cls="px-2 py-1 text-xs font-mono"),
                        Td(
                            Span(f.get("storey") or "—", cls="block text-xs"),
                            Span(f.get("space") or "", cls="block text-xs text-muted-foreground"),
                            cls="px-2 py-1",
                        ),
                        Td(_fmt_actual(f.get("actual")), cls="px-2 py-1 text-xs font-mono"),
                        Td(f.get("reason", ""), cls="px-2 py-1 text-xs text-red-700"),
                        Td(
                            Span(
                                (f.get("guid") or "")[:12],
                                cls="text-xs text-muted-foreground font-mono",
                            ),
                            cls="px-2 py-1",
                        ),
                        Td(_view_in_3d_link(f), cls="px-2 py-1"),
                    )
                    for f in failures[:20]
                ]
                fail_detail = Details(
                    Summary(
                        f"{len(failures)} failing element(s) — click to expand",
                        cls="text-xs text-red-600 cursor-pointer font-semibold mt-1",
                    ),
                    Div(
                        Table(
                            Thead(
                                Tr(
                                    Th("Element", cls="px-2 py-1 text-xs bg-muted"),
                                    Th("Floor / Room", cls="px-2 py-1 text-xs bg-muted"),
                                    Th("Actual value", cls="px-2 py-1 text-xs bg-muted"),
                                    Th("Reason", cls="px-2 py-1 text-xs bg-muted"),
                                    Th("GUID", cls="px-2 py-1 text-xs bg-muted"),
                                    Th("", cls="px-2 py-1 text-xs bg-muted"),
                                )
                            ),
                            Tbody(*fail_rows),
                            cls="w-full text-xs border rounded mt-1",
                        ),
                        cls="max-h-48 overflow-y-auto",
                    ),
                )

            # Collapsible missing-data details
            missing_elements = r.get("missing_elements", [])
            missing_detail = ""
            if missing_elements:
                missing_rows = [
                    Tr(
                        Td(m.get("element_name", "")[:35], cls="px-2 py-1 text-xs font-mono"),
                        Td(m.get("storey") or "—", cls="px-2 py-1 text-xs"),
                        Td(m.get("space") or "—", cls="px-2 py-1 text-xs text-muted-foreground"),
                        Td(
                            m.get("guid", "")[:16],
                            cls="px-2 py-1 text-xs text-muted-foreground font-mono",
                        ),
                    )
                    for m in missing_elements[:20]
                ]
                overflow_note = (
                    Div(
                        f"... and {len(missing_elements) - 20} more elements",
                        cls="text-xs text-muted-foreground px-2 py-1",
                    )
                    if len(missing_elements) > 20
                    else ""
                )
                missing_detail = Details(
                    Summary(
                        f"{len(missing_elements)} element(s) missing this property — click to expand",
                        cls="text-xs text-yellow-700 cursor-pointer font-semibold mt-1",
                    ),
                    Div(
                        Table(
                            Thead(
                                Tr(
                                    Th("Element", cls="px-2 py-1 text-xs bg-muted"),
                                    Th("Floor", cls="px-2 py-1 text-xs bg-muted"),
                                    Th("Room", cls="px-2 py-1 text-xs bg-muted"),
                                    Th("GUID", cls="px-2 py-1 text-xs bg-muted"),
                                )
                            ),
                            Tbody(*missing_rows),
                            cls="w-full text-xs border rounded mt-1",
                        ),
                        overflow_note,
                        cls="max-h-48 overflow-y-auto",
                    ),
                )

            rows.append(
                Tr(
                    Td(r.get("rule_ref", "—"), cls="px-3 py-2 text-xs font-mono"),
                    Td(
                        Div(
                            (r.get("rule_desc", "") or "")[:70]
                            + ("..." if len(r.get("rule_desc", "") or "") > 70 else ""),
                            fail_detail,
                            missing_detail,
                        ),
                        cls="px-3 py-2 text-xs",
                    ),
                    Td(r.get("property_name", ""), cls="px-3 py-2 text-xs font-mono"),
                    Td(rule_str, cls="px-3 py-2 text-xs font-mono"),
                    Td(
                        Span(
                            status,
                            cls=f"inline-block px-1.5 py-0.5 rounded text-xs font-semibold {s_cls}",
                        ),
                        cls="px-3 py-2",
                    ),
                    Td(counts, cls="px-3 py-2"),
                    cls="border-b border-muted last:border-0",
                )
            )

    csv_btn = A(
        "Download CSV",
        href="/reports/compliance-csv",
        cls="inline-block px-3 py-1.5 rounded text-xs font-medium bg-slate-800 text-white hover:bg-slate-600 mt-3",
    )
    bcf_btn = A(
        "Download BCF (Revit / Solibri)",
        href="/reports/compliance-bcf",
        cls="inline-block px-3 py-1.5 rounded text-xs font-medium bg-blue-700 text-white hover:bg-blue-600 mt-3 ml-2",
    )
    view_3d_btn = (
        A(
            "View failures in 3D",
            href=f"/viewer?project_id={project_id}&bcf_artifact_id={bcf_artifact_id}",
            target="_blank",
            cls="inline-block px-3 py-1.5 rounded text-xs font-medium bg-emerald-700 text-white hover:bg-emerald-600 mt-3 ml-2",
        )
        if project_id is not None and bcf_artifact_id is not None
        else ""
    )

    return _collapsible_card(
        title,
        summary_bar,
        _coverage_stat_line(summary),
        Div(
            Table(Thead(Tr(*header_cells)), Tbody(*rows), cls="w-full text-sm"),
            cls="overflow-auto border rounded-md",
        ),
        Div(csv_btn, bcf_btn, view_3d_btn),
    )


_ELEM_LABELS = {
    "IfcDoor": "Doors",
    "IfcWindow": "Windows",
    "IfcWall": "Walls",
    "IfcCurtainWall": "Curtain Walls",
    "IfcSlab": "Slabs",
    "IfcRoof": "Roofs",
    "IfcCovering": "Ceilings",
    "IfcStairFlight": "Stair Flights",
    "IfcRamp": "Ramps",
    "IfcRampFlight": "Ramp Flights",
    "IfcRailing": "Railings",
    "IfcColumn": "Columns",
    "IfcBeam": "Beams",
    "IfcMember": "Structural Members",
    "IfcSanitaryTerminal": "Fixtures",
    "IfcAlarm": "Alarms",
    "IfcSensor": "Sensors",
    "IfcFurnishingElement": "Furniture",
}


def _collapsible_card(title: str, *content, subtitle: str = "", open: bool = True):
    """Top-level collapsible result card with a clickable header."""
    return Details(
        Summary(
            Div(
                Span(title, cls="font-semibold text-base"),
                Span(subtitle, cls="text-xs text-muted-foreground ml-2") if subtitle else "",
                cls="flex items-center gap-1",
            ),
            cls="px-6 py-4 border-b bg-muted/30 cursor-pointer select-none list-none",
        ),
        Div(*content, cls="px-6 py-5 space-y-4 bg-card"),
        cls="border rounded-lg shadow-sm overflow-hidden",
        open=open,
    )


def _collapsible_section(title: str, *content, badge: str = "", open: bool = True, extra=None):
    """Collapsible sub-section inside a card (smaller heading).

    `extra` is an optional control (e.g. a select) shown at the right edge of
    the header — wrapped with a click-stop so interacting with it doesn't
    toggle the surrounding <details>/<summary>.
    """
    return Details(
        Summary(
            Div(
                Span(title, cls="text-sm font-semibold"),
                Span(
                    badge,
                    cls="ml-2 text-xs px-2 py-0.5 rounded-full bg-muted text-muted-foreground",
                )
                if badge
                else "",
                Div(extra, cls="ml-auto", onclick="event.stopPropagation()") if extra else "",
                cls="flex items-center",
            ),
            cls="cursor-pointer select-none py-1 list-none",
        ),
        Div(*content, cls="mt-2"),
        open=open,
    )


def _building_summary_card(summary: dict):
    """Render a Building Overview card from extract_building_summary() output."""
    if not summary:
        return ""

    storey_count = summary.get("storey_count", 0)
    room_count = summary.get("room_count", 0)
    gfa = summary.get("total_gfa_m2", 0.0)
    ext_doors = summary.get("external_door_count", 0)
    element_counts = summary.get("element_counts", {})
    fixture_counts = summary.get("fixture_counts", {})
    alarm_counts = summary.get("alarm_counts", {})
    floor_heights = summary.get("floor_heights", [])
    rooms_per_storey = summary.get("rooms_per_storey", {})
    storeys = summary.get("storeys", [])
    unplaced = summary.get("unplaced_rooms", [])
    unnamed = summary.get("unnamed_elements", [])

    # ── Stat strip ────────────────────────────────────────────────────────────
    def _stat(label, value):
        return Div(
            Span(str(value), cls="text-2xl font-bold block"),
            Span(label, cls="text-xs text-muted-foreground"),
            cls="text-center px-4 py-3 bg-muted rounded-lg",
        )

    stat_strip = Div(
        _stat("Storeys", storey_count),
        _stat("Rooms / Spaces", room_count),
        _stat("GFA m²", f"{gfa:,.1f}" if gfa else "—"),
        _stat("Exit Doors", ext_doors),
        cls="grid grid-cols-4 gap-3 mb-5",
    )

    # ── Floor breakdown table ─────────────────────────────────────────────────
    floor_table = ""
    if storeys:
        fh_by_from = {h["from"]: h["height_mm"] for h in floor_heights}
        floor_rows = []
        for s in storeys:
            name = s["name"]
            ri = rooms_per_storey.get(name, {})
            r_cnt = ri.get("count", 0)
            r_area = ri.get("total_area_m2", 0.0)
            h_mm = fh_by_from.get(name)
            floor_rows.append(
                Tr(
                    Td(name, cls="px-3 py-2 text-sm font-medium"),
                    Td(
                        (f"{h_mm / 1000:.2f} m" if h_mm >= 1000 else f"{h_mm:,} mm")
                        if h_mm
                        else "—",
                        cls="px-3 py-2 text-sm font-mono",
                    ),
                    Td(str(r_cnt) if r_cnt else "—", cls="px-3 py-2 text-sm text-center"),
                    Td(
                        f"{r_area:,.1f}" if r_area else "—",
                        cls="px-3 py-2 text-sm font-mono text-right",
                    ),
                    cls="border-b border-muted last:border-0",
                )
            )
        floor_table = Div(
            H3("Floor Breakdown", cls="text-sm font-semibold mb-2"),
            Div(
                Table(
                    Thead(
                        Tr(
                            Th(
                                "Storey",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Floor-to-Floor",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Rooms",
                                cls="px-3 py-2 text-center text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Area m²",
                                cls="px-3 py-2 text-right text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                        )
                    ),
                    Tbody(*floor_rows),
                    cls="w-full text-sm",
                ),
                cls="overflow-auto border rounded-md mb-5",
            ),
        )

    # ── Element count badges ──────────────────────────────────────────────────
    elem_badges = ""
    if element_counts:
        elem_badges = Div(
            H3("Elements Found", cls="text-sm font-semibold mb-2"),
            Div(
                *[
                    Span(
                        f"{_ELEM_LABELS.get(k, k)}: {v}",
                        cls="inline-block px-2 py-1 rounded-full text-xs bg-blue-50 text-blue-800 border border-blue-200 font-medium",
                    )
                    for k, v in sorted(
                        element_counts.items(), key=lambda x: _ELEM_LABELS.get(x[0], x[0])
                    )
                ],
                cls="flex flex-wrap gap-2 mb-5",
            ),
        )

    # ── Plumbing fixture badges ───────────────────────────────────────────────
    fixture_badges = ""
    if fixture_counts:
        fixture_badges = Div(
            H3("Plumbing Fixtures", cls="text-sm font-semibold mb-2"),
            Div(
                *[
                    Span(
                        f"{k}: {v}",
                        cls="inline-block px-2 py-1 rounded-full text-xs bg-cyan-50 text-cyan-800 border border-cyan-200 font-medium",
                    )
                    for k, v in sorted(fixture_counts.items())
                ],
                cls="flex flex-wrap gap-2 mb-5",
            ),
        )

    # ── Alarm badges ─────────────────────────────────────────────────────────
    alarm_badges = ""
    if alarm_counts:
        alarm_badges = Div(
            H3("Fire / CO Alarms", cls="text-sm font-semibold mb-2"),
            Div(
                *[
                    Span(
                        f"{k}: {v}",
                        cls="inline-block px-2 py-1 rounded-full text-xs bg-red-50 text-red-800 border border-red-200 font-medium",
                    )
                    for k, v in sorted(alarm_counts.items())
                ],
                cls="flex flex-wrap gap-2 mb-5",
            ),
        )

    # ── QA issues ─────────────────────────────────────────────────────────────
    qa_items = []
    if unplaced:
        qa_items.append(
            P(
                f"⚠ {len(unplaced)} unplaced room(s) — not assigned to any storey",
                cls="text-xs text-yellow-700",
            )
        )
    for u in unnamed:
        qa_items.append(
            P(
                f"⚠ {u['count']} {_ELEM_LABELS.get(u['type'], u['type'])} element(s) missing Name property",
                cls="text-xs text-yellow-700",
            )
        )
    qa_block = ""
    if qa_items:
        qa_block = Div(
            H3("Model QA", cls="text-sm font-semibold mb-2"),
            Div(*qa_items, cls="space-y-1 p-3 bg-yellow-50 rounded-md border border-yellow-200"),
        )

    return _collapsible_card(
        "Building Overview",
        stat_strip,
        _collapsible_section("Floor Breakdown", floor_table) if floor_table else "",
        _collapsible_section("Elements Found", elem_badges) if elem_badges else "",
        _collapsible_section("Plumbing Fixtures", fixture_badges) if fixture_badges else "",
        _collapsible_section("Fire / CO Alarms", alarm_badges) if alarm_badges else "",
        qa_block,
    )


def _spatial_checks_card(spatial: dict):
    """Render Tier 2 spatial compliance results — daylight ratios and fire separation."""
    if not spatial:
        return ""

    has_boundaries = spatial.get("has_boundaries", False)
    warnings = spatial.get("warnings", [])
    daylight = spatial.get("daylight", [])
    fire_sep = spatial.get("fire_separation", [])
    garage_sep = spatial.get("garage_separation", {})
    code_ctx = _code_rule_context()

    if not has_boundaries:
        msg = (
            warnings[0]
            if warnings
            else "No IfcRelSpaceBoundary data — re-export with Space Boundaries enabled."
        )
        return _collapsible_card(
            "Spatial Compliance — Tier 2",
            Span(msg, cls="text-sm text-yellow-700"),
        )

    warning_items = [Span(w, cls="block text-xs text-yellow-700") for w in warnings]

    # ── Daylight ratio table ──────────────────────────────────────────────────
    daylight_section = ""
    if daylight:
        daylight_ref = next(
            (_result_ref(r) for r in daylight if _result_ref(r)),
            code_ctx["daylight_ref"],
        )
        d_pass = sum(1 for r in daylight if r["passes"])
        d_fail = len(daylight) - d_pass
        rate_cls = "bg-green-100 text-green-800" if d_fail == 0 else "bg-red-100 text-red-800"
        d_rows = []
        for r in sorted(daylight, key=lambda x: x["passes"]):
            status_cls = (
                "text-green-700 font-semibold" if r["passes"] else "text-red-700 font-semibold"
            )
            d_rows.append(
                Tr(
                    Td(r.get("storey_name") or "—", cls="px-3 py-2 text-xs text-muted-foreground"),
                    Td(r["space_name"][:35], cls="px-3 py-2 text-xs"),
                    Td(f"{r['floor_area_m2']:.1f}", cls="px-3 py-2 text-xs font-mono text-right"),
                    Td(
                        f"{r['total_window_area_m2']:.2f}",
                        cls="px-3 py-2 text-xs font-mono text-right",
                    ),
                    Td(f"{r['daylight_ratio']:.3f}", cls="px-3 py-2 text-xs font-mono text-right"),
                    Td(
                        "✓ Pass" if r["passes"] else "✗ Fail", cls=f"px-3 py-2 text-xs {status_cls}"
                    ),
                    cls="border-b border-muted last:border-0",
                )
            )
        daylight_section = _collapsible_section(
            _section_title("Daylight Ratio", daylight_ref),
            Div(
                Table(
                    Thead(
                        Tr(
                            Th(
                                "Floor",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Room",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Floor m²",
                                cls="px-3 py-2 text-right text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Window m²",
                                cls="px-3 py-2 text-right text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Ratio",
                                cls="px-3 py-2 text-right text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Status",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                        )
                    ),
                    Tbody(*d_rows),
                    cls="w-full text-sm",
                ),
                cls="overflow-auto border rounded-md",
            ),
            badge=f"{d_pass}/{len(daylight)} pass",
        )

    # ── Fire separation table ─────────────────────────────────────────────────
    fire_section = ""
    if fire_sep:
        fire_ref = next(
            (_result_ref(r) for r in fire_sep if _result_ref(r)),
            code_ctx["fire_ref"],
        )
        f_pass = sum(1 for r in fire_sep if r["passes"])
        f_fail = len(fire_sep) - f_pass
        f_rate_cls = "bg-green-100 text-green-800" if f_fail == 0 else "bg-red-100 text-red-800"
        f_rows = []
        for r in sorted(fire_sep, key=lambda x: x["passes"]):
            rating_txt = r["fire_rating_raw"] or "⚠ Not declared"
            rating_cls = (
                "text-red-700"
                if r["missing_rating"]
                else ("text-green-700" if r["passes"] else "text-orange-700")
            )
            spaces_txt = ", ".join(r["adjacent_spaces"][:2])
            if len(r["adjacent_spaces"]) > 2:
                spaces_txt += f" +{len(r['adjacent_spaces']) - 2}"
            f_rows.append(
                Tr(
                    Td(r["wall_name"][:35], cls="px-3 py-2 text-xs font-mono"),
                    Td(spaces_txt[:50], cls="px-3 py-2 text-xs"),
                    Td(rating_txt, cls=f"px-3 py-2 text-xs font-mono {rating_cls}"),
                    Td(
                        "✓ Pass" if r["passes"] else "✗ Fail", cls=f"px-3 py-2 text-xs {rating_cls}"
                    ),
                    cls="border-b border-muted last:border-0",
                )
            )
        fire_section = _collapsible_section(
            _section_title("Fire Separation", fire_ref),
            Div(
                Table(
                    Thead(
                        Tr(
                            Th(
                                "Wall",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Between Spaces",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Fire Rating",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Status",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                        )
                    ),
                    Tbody(*f_rows),
                    cls="w-full text-sm",
                ),
                cls="overflow-auto border rounded-md",
            ),
            badge=f"{f_pass}/{len(fire_sep)} pass",
        )

    # ── Garage separation section ─────────────────────────────────────────────
    garage_section = ""
    if garage_sep:
        g_results = garage_sep.get("results", [])
        g_warnings = garage_sep.get("warnings", [])
        g_found = garage_sep.get("garage_spaces_found", 0)
        garage_ref = next(
            (_result_ref(r) for r in g_results if _result_ref(r)),
            code_ctx["garage_ref"],
        )

        if g_results:
            g_pass = sum(1 for r in g_results if r["passes"])
            g_rows = []
            for r in sorted(g_results, key=lambda x: x["passes"]):
                rating_txt = r["fire_rating_raw"] or "⚠ Not declared"
                rating_cls = (
                    "text-red-700"
                    if r["missing_rating"]
                    else ("text-green-700" if r["passes"] else "text-orange-700")
                )
                req_txt = f"≥ {r['required_min']} min"
                g_rows.append(
                    Tr(
                        Td(r["element_type"], cls="px-3 py-2 text-xs font-semibold"),
                        Td(r["element_name"][:35], cls="px-3 py-2 text-xs font-mono"),
                        Td(r["garage_space"][:25], cls="px-3 py-2 text-xs"),
                        Td(r["adjacent_space"][:25], cls="px-3 py-2 text-xs"),
                        Td(rating_txt, cls=f"px-3 py-2 text-xs font-mono {rating_cls}"),
                        Td(req_txt, cls="px-3 py-2 text-xs font-mono"),
                        Td(
                            "✓ Pass" if r["passes"] else "✗ Fail",
                            cls=f"px-3 py-2 text-xs {rating_cls}",
                        ),
                        cls="border-b border-muted last:border-0",
                    )
                )
            garage_content = Div(
                Table(
                    Thead(
                        Tr(
                            Th(
                                "Type",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Element",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Garage Space",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Adjacent Space",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Fire Rating",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Required",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Status",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                        )
                    ),
                    Tbody(*g_rows),
                    cls="w-full text-sm",
                ),
                cls="overflow-auto border rounded-md",
            )
            badge = f"{g_pass}/{len(g_results)} pass"
        else:
            garage_content = Div(
                *[Span(w, cls="block text-xs text-yellow-700") for w in g_warnings],
                cls="space-y-1",
            )
            badge = f"{g_found} garage(s)"

        garage_section = _collapsible_section(
            _section_title("Garage Separation", garage_ref),
            garage_content,
            badge=badge,
        )

    subtitle = (
        f"{spatial.get('space_count', 0)} spaces · {spatial.get('party_wall_count', 0)} party walls"
    )
    return _collapsible_card(
        "Spatial Compliance — Tier 2",
        Div(*[Span(w, cls="block text-xs text-yellow-700") for w in warnings], cls="space-y-1")
        if warnings
        else "",
        daylight_section,
        fire_section,
        garage_section,
        subtitle=subtitle,
    )


def _egress_checks_card(egress: dict):
    """Render Tier 3 egress compliance results — exit count and travel distances."""
    if not egress:
        return ""

    exit_data = egress.get("exit_count", {})
    travel = egress.get("travel_distance", [])
    has_graph = egress.get("has_graph", False)
    warnings = egress.get("warnings", [])

    total_exits = exit_data.get("total_exterior_doors", 0)
    exit_results = exit_data.get("results", [])
    exit_warnings = exit_data.get("warnings", [])
    all_warnings = exit_warnings + warnings
    code_ctx = _code_rule_context()
    exit_ref = next(
        (_result_ref(r) for r in exit_results if _result_ref(r)),
        code_ctx["exit_ref"],
    )

    # ── Exit count section ────────────────────────────────────────────────────
    exit_rows = []
    for r in exit_results:
        status_cls = "text-green-700 font-semibold" if r["passes"] else "text-red-700 font-semibold"
        exit_rows.append(
            Tr(
                Td(r["storey"], cls="px-3 py-2 text-xs"),
                Td(str(r["exit_count"]), cls="px-3 py-2 text-xs font-mono text-center"),
                Td(str(r["required_min"]), cls="px-3 py-2 text-xs font-mono text-center"),
                Td("✓ Pass" if r["passes"] else "✗ Fail", cls=f"px-3 py-2 text-xs {status_cls}"),
                cls="border-b border-muted last:border-0",
            )
        )

    exit_section = _collapsible_section(
        _section_title("Exit Count", exit_ref),
        Div(
            P(f"Total exterior doors detected: {total_exits}", cls="text-sm mb-2"),
            Div(
                Table(
                    Thead(
                        Tr(
                            Th(
                                "Storey",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Exits",
                                cls="px-3 py-2 text-center text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Required",
                                cls="px-3 py-2 text-center text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Status",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                        )
                    ),
                    Tbody(*exit_rows)
                    if exit_rows
                    else Tbody(
                        Tr(
                            Td(
                                "No exterior doors found — tag doors as IsExternal=True",
                                cls="px-3 py-2 text-xs text-yellow-700",
                                colspan="4",
                            )
                        )
                    ),
                    cls="w-full text-sm",
                ),
                cls="overflow-auto border rounded-md",
            ),
        )
        if exit_results
        else P(
            "No exterior doors found — tag doors as IsExternal=True in the authoring tool.",
            cls="text-sm text-yellow-700",
        ),
        badge=f"{total_exits} exit(s)",
    )

    # ── Travel distance section ───────────────────────────────────────────────
    travel_section = ""
    if travel:
        travel_ref = next(
            (_result_ref(r) for r in travel if _result_ref(r)),
            code_ctx["travel_ref"],
        )
        td_pass = sum(1 for r in travel if r["passes"])
        td_fail = len(travel) - td_pass
        td_rows = []
        for r in sorted(travel, key=lambda x: (x["passes"], -(x.get("travel_distance_m") or 0))):
            dist_txt = (
                f"{r['travel_distance_m']:.1f} m"
                if r.get("travel_distance_m") is not None
                else "No path"
            )
            status_cls = (
                "text-green-700 font-semibold" if r["passes"] else "text-red-700 font-semibold"
            )
            td_rows.append(
                Tr(
                    Td(r.get("storey_name") or "—", cls="px-3 py-2 text-xs text-muted-foreground"),
                    Td(r["space_name"][:35], cls="px-3 py-2 text-xs"),
                    Td(dist_txt, cls="px-3 py-2 text-xs font-mono text-right"),
                    Td(r.get("nearest_exit") or "—", cls="px-3 py-2 text-xs"),
                    Td(
                        "✓ Pass"
                        if r["passes"]
                        else ("✗ No path" if r.get("no_path") else "✗ Exceeds"),
                        cls=f"px-3 py-2 text-xs {status_cls}",
                    ),
                    cls="border-b border-muted last:border-0",
                )
            )
        travel_section = _collapsible_section(
            _section_title("Travel Distance", travel_ref),
            Div(
                Table(
                    Thead(
                        Tr(
                            Th(
                                "Floor",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Room",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Distance",
                                cls="px-3 py-2 text-right text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Nearest Exit",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                            Th(
                                "Status",
                                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                            ),
                        )
                    ),
                    Tbody(*td_rows),
                    cls="w-full text-sm",
                ),
                cls="overflow-auto border rounded-md",
            ),
            badge=f"{td_pass}/{len(travel)} pass",
        )
    elif not has_graph:
        travel_ref = code_ctx["travel_ref"]
        travel_section = _collapsible_section(
            _section_title("Travel Distance", travel_ref),
            P(
                "Space boundary data required. Re-export the IFC model with "
                "Space Boundaries enabled to activate this check.",
                cls="text-sm text-yellow-700",
            ),
        )

    warning_items = [Span(w, cls="block text-xs text-yellow-700") for w in all_warnings]

    subtitle = f"{total_exits} exit(s) · {len(travel)} rooms checked"
    return _collapsible_card(
        "Egress Compliance — Tier 3",
        Div(*warning_items, cls="space-y-1") if warning_items else "",
        exit_section,
        travel_section,
        subtitle=subtitle,
    )


# Module-level cache for CSV export (populated in analysis_run_post)
_last_compliance_results: list[dict] = []

# Module-level cache for the Simple Analysis per-category Excel export
# (populated in analysis_simple_post)
_last_simple_compliance: list[dict] = []

_SIMPLE_CATEGORY_TARGETS: dict[str, list[str]] = {
    "windows": ["IfcWindow"],
    "doors": ["IfcDoor"],
    "stairs": ["IfcStairFlight", "IfcRailing"],
    "ramps": ["IfcRamp", "IfcRampFlight"],
    "washrooms": ["IfcSanitaryTerminal"],
    "fire": ["IfcAlarm"],
}


# ══════════════════════════════════════════════════════════════════════════════
# Simple Analysis — domain-based compliance card helpers
# ══════════════════════════════════════════════════════════════════════════════


def _fmt_val_s(v) -> str:
    """Format a numeric compliance value for display."""
    if isinstance(v, float):
        return f"{v:,.2f}" if v >= 1 else f"{v:.4f}"
    return str(v) if v is not None else "—"


def _fmt_duration(seconds: float) -> str:
    """Format a run duration for display — 's' under a minute, 'm s' beyond."""
    if seconds < 60:
        return f"{seconds:.1f}s"
    minutes, rest = divmod(seconds, 60)
    return f"{int(minutes)}m {rest:.0f}s"


def _pass_rate_pill(summary: dict, total_key: str = "total_rules"):
    """Colour-coded 'N% pass rate' pill, shared by the MEP rule compliance
    card and the ARCH results page. Returns "" when there's nothing to rate
    (no rules evaluated) rather than a misleading "0% pass rate"."""
    total = summary.get(total_key, 0)
    if not total:
        return ""
    rate = summary.get("pass_rate", 0)
    rate_cls = (
        "bg-green-100 text-green-800"
        if rate >= 80
        else "bg-yellow-100 text-yellow-800"
        if rate >= 50
        else "bg-red-100 text-red-800"
    )
    return Span(
        f"{rate:.0f}% pass rate",
        cls=f"inline-block px-3 py-1 rounded-full text-sm font-semibold {rate_cls} mr-3",
    )


def _coverage_stat_line(summary: dict):
    """Duration + elements-checked evidence line, shared by the MEP rule
    compliance card and the ARCH results page — the concrete "how long, how
    much" numbers behind an automated-vs-manual pitch (no meaningful line to
    show without at least one of the two)."""
    duration = summary.get("duration_seconds")
    elements = summary.get("unique_elements_evaluated", 0)
    rules_with_elements = summary.get("rules_with_elements", 0)
    if duration is None and not elements:
        return ""
    parts = []
    if duration is not None:
        parts.append(Span(f"⏱ {_fmt_duration(duration)}", cls="text-xs text-muted-foreground mr-3"))
    if elements:
        parts.append(
            Span(
                f"🔍 {elements} element(s) checked across {rules_with_elements} applicable "
                f"rule(s) — every match evaluated, no sampling",
                cls="text-xs text-muted-foreground",
            )
        )
    return Div(*parts, cls="flex flex-wrap items-center gap-1 mb-3")


def _domain_badge(rule_results: list) -> tuple:
    """Return (label, tailwind_cls) from a list of Module 4 rule result dicts."""
    if not rule_results:
        return "N/A", "bg-gray-100 text-gray-500"
    statuses = [r.get("status", "NO_ELEMENTS") for r in rule_results]
    real = [s for s in statuses if s != "NO_ELEMENTS"]
    if not real:
        return "N/A", "bg-gray-100 text-gray-500"
    n_fail = real.count("FAIL")
    if n_fail:
        return f"{n_fail} rule(s) failed", "bg-red-100 text-red-800"
    if any(s in ("MISSING_DATA", "PARTIAL") for s in real):
        return "Missing data", "bg-yellow-100 text-yellow-800"
    return "All pass", "bg-green-100 text-green-800"


def _simple_full_elem_tbl(all_elements: list, rule: dict):
    """Scrollable element table for one rule — every evaluated element (pass,
    fail, and missing), sorted so problems surface first."""
    if not all_elements:
        return P("No elements.", cls="text-xs text-muted-foreground")

    required = _rule_required_text(rule)
    unit = rule.get("unit", "") or ""
    order = {"FAIL": 0, "MISSING": 1, "PASS": 2}
    ordered = sorted(all_elements, key=lambda el: order.get(el.get("status"), 3))

    rows = []
    for el in ordered[:50]:
        status = el.get("status")
        actual = el.get("actual")
        actual_txt = _fmt_val_s(actual) + (f" {unit}" if unit and actual is not None else "")
        storey = el.get("storey") or "—"
        space = el.get("space") or ""
        guid = (el.get("guid") or "")[:14]
        name = (el.get("element_name") or "—")[:32]
        if status == "FAIL":
            status_txt, status_cls, row_cls = (
                el.get("reason") or "fail",
                "text-red-700 font-semibold",
                "bg-red-50",
            )
        elif status == "MISSING":
            status_txt, status_cls, row_cls = (
                "missing",
                "text-yellow-700 font-semibold",
                "bg-yellow-50",
            )
        else:
            status_txt, status_cls, row_cls = "✓ pass", "text-green-700", ""
        rows.append(
            Tr(
                Td(Span(name, cls="text-xs font-mono"), cls="px-3 py-2"),
                Td(
                    Span(storey, cls="text-xs block"),
                    Span(space, cls="text-xs text-muted-foreground block")
                    if space and space != "—"
                    else "",
                    cls="px-3 py-2",
                ),
                Td(Span(guid, cls="text-xs font-mono text-muted-foreground"), cls="px-3 py-2"),
                Td(actual_txt, cls="px-3 py-2 text-xs font-mono"),
                Td(required, cls="px-3 py-2 text-xs text-muted-foreground"),
                Td(Span(status_txt, cls=f"text-xs {status_cls}"), cls="px-3 py-2"),
                cls=f"border-b border-muted last:border-0 {row_cls}",
            )
        )
    if len(ordered) > 50:
        rows.append(
            Tr(
                Td(
                    f"… and {len(ordered) - 50} more",
                    cls="px-3 py-2 text-xs text-muted-foreground italic",
                    colspan="6",
                )
            )
        )
    return Div(
        Table(
            Thead(
                Tr(
                    *[
                        Th(
                            h,
                            cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                        )
                        for h in ("Element", "Floor / Room", "GUID", "Actual", "Required", "Status")
                    ]
                )
            ),
            Tbody(*rows),
            cls="w-full text-xs",
        ),
        cls="overflow-auto border rounded-md max-h-64",
    )


def _rule_required_text(rule: dict) -> str:
    """Human-readable requirement string for one Module 4 rule result."""
    operator = rule.get("operator", "")
    check_val = rule.get("check_value")
    unit = rule.get("unit", "") or ""
    if operator == ">=":
        return f"≥ {check_val} {unit}".strip()
    if operator == "<=":
        return f"≤ {check_val} {unit}".strip()
    if operator == "between":
        return f"{rule.get('value_min')}–{rule.get('value_max')} {unit}".strip()
    if operator == "exists":
        return "must be present"
    if operator == "not_exists":
        return "must not be present"
    return f"{operator} {check_val} {unit}".strip() if check_val is not None else "—"


def _pivot_category_elements(rule_results: list) -> tuple[list, list]:
    """
    Pivot a category's per-rule Module 4 results (e.g. all IfcDoor rules) into
    one row per physical element with one actual/required pair per rule.

    Returns (columns, rows):
      columns — the rule_results entries actually used (NO_ELEMENTS ones dropped)
      rows    — [{"name","storey","space","guid","cells": {rule_key: entry|None}, "issues": [str]}]
    """
    columns = [r for r in rule_results if r.get("status") != "NO_ELEMENTS"]
    if not columns:
        return [], []

    identities: dict[str, dict] = {}
    cells: dict[str, dict[str, dict]] = {}

    for rule in columns:
        rule_key = rule.get("rule_ref") or rule.get("property_name") or str(id(rule))
        for el in rule.get("all_elements", []):
            guid = el.get("guid") or el.get("element_name") or ""
            if not guid:
                continue
            if guid not in identities:
                identities[guid] = {
                    "name": el.get("element_name") or "—",
                    "storey": el.get("storey") or "—",
                    "space": el.get("space") or "—",
                    "guid": guid,
                    "position_mm": None,
                }
            # Different rules can evaluate the same element; geometry only
            # needs resolving once, so keep the first position_mm any of
            # them managed to resolve rather than overwriting with a later
            # None (e.g. a rule whose element pass never hit Module 2's
            # geometry fallback).
            if identities[guid]["position_mm"] is None and el.get("position_mm") is not None:
                identities[guid]["position_mm"] = el["position_mm"]
            cells.setdefault(guid, {})[rule_key] = el

    rows = []
    for guid, identity in identities.items():
        row_cells = {}
        issues = []
        for rule in columns:
            rule_key = rule.get("rule_ref") or rule.get("property_name") or str(id(rule))
            entry = cells.get(guid, {}).get(rule_key)
            row_cells[rule_key] = entry
            if entry and entry.get("status") != "PASS":
                label = rule.get("property_name") or rule_key
                issues.append(f"{label}: {entry.get('reason') or 'missing'}")
        rows.append({**identity, "cells": row_cells, "issues": issues})

    rows.sort(key=lambda r: (len(r["issues"]) == 0, r["name"]))
    return columns, rows


# Pixel widths + left offsets for the three frozen identity columns in the
# summary table, so they stay visible while scrolling through property columns.
_STICKY_COL_WIDTHS = [150, 110, 120]
_STICKY_COL_OFFSETS = [0, 150, 260]


def _simple_summary_table(category: str, rule_results: list):
    """Always-on per-element summary: one row per element tested, one
    actual/required column pair per rule, and an aggregated Issues column.
    Rows/cells with a failing or missing property are highlighted red/yellow.
    Full (untruncated) data is available via the Download Excel / PDF buttons."""
    columns, rows = _pivot_category_elements(rule_results)

    if not columns:
        return ""

    _bcf_artifact = (
        _report_artifact_service.latest_bcf(_last_summary_project_id)
        if _last_summary_project_id is not None
        else None
    )
    download_btns = Div(
        A(
            "Download Excel",
            href=f"/reports/simple-summary-xlsx/{category}",
            cls="inline-block px-3 py-1.5 rounded text-xs font-medium bg-slate-800 text-white hover:bg-slate-600",
        ),
        A(
            "Download PDF",
            href=f"/reports/simple-summary-pdf/{category}",
            cls="inline-block px-3 py-1.5 rounded text-xs font-medium bg-slate-800 text-white hover:bg-slate-600",
        ),
        A(
            "View in 3D",
            href=f"/viewer?project_id={_last_summary_project_id}&bcf_artifact_id={_bcf_artifact['id']}",
            target="_blank",
            cls="inline-block px-3 py-1.5 rounded text-xs font-medium bg-emerald-700 text-white hover:bg-emerald-600",
        )
        if _bcf_artifact
        else "",
        cls="flex gap-2",
    )

    if not rows:
        return Div(
            P(
                "No elements found in the model for this category.",
                cls="text-xs text-muted-foreground italic mb-2",
            ),
            cls="mb-4",
        )

    id_headers = ["Element", "Floor / Room", "GUID"]

    def _sticky(i: int) -> str:
        return f"sticky left-[{_STICKY_COL_OFFSETS[i]}px] z-20"

    row1 = [
        Th(
            h,
            rowspan="2",
            cls=f"px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted align-bottom {_sticky(i)}",
        )
        for i, h in enumerate(id_headers)
    ]
    row2 = []
    for rule in columns:
        label = rule.get("property_name") or rule.get("rule_ref") or "?"
        row1.append(
            Th(
                Span(label, cls="block"),
                Span(
                    rule.get("rule_ref", ""),
                    cls="block text-[10px] font-normal text-muted-foreground",
                ),
                colspan="2",
                cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted border-l-2 border-slate-300",
            )
        )
        row2.append(
            Th(
                "Actual",
                cls="px-3 py-2 text-left text-[10px] font-medium text-muted-foreground bg-muted border-l-2 border-slate-300",
            )
        )
        row2.append(
            Th(
                "Required",
                cls="px-3 py-2 text-left text-[10px] font-medium text-muted-foreground bg-muted",
            )
        )
    row1.append(
        Th(
            "Issues",
            rowspan="2",
            cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted align-bottom",
        )
    )

    body_rows = []
    display_rows = rows[:50]
    for r in display_rows:
        row_has_issue = bool(r["issues"])
        row_bg = "bg-red-50/60" if row_has_issue else "bg-white"
        id_cells = [
            Td(
                Span(r["name"][:32], cls="text-xs font-mono"),
                cls=f"px-3 py-2 {_sticky(0)} {row_bg}",
                style=f"width:{_STICKY_COL_WIDTHS[0]}px",
            ),
            Td(
                Span(r["storey"], cls="text-xs block"),
                Span(r["space"], cls="text-xs text-muted-foreground block")
                if r["space"] not in ("", "—")
                else "",
                cls=f"px-3 py-2 {_sticky(1)} {row_bg}",
                style=f"width:{_STICKY_COL_WIDTHS[1]}px",
            ),
            Td(
                Span(r["guid"][:14], cls="text-xs font-mono text-muted-foreground"),
                cls=f"px-3 py-2 {_sticky(2)} {row_bg}",
                style=f"width:{_STICKY_COL_WIDTHS[2]}px",
            ),
        ]
        tds = list(id_cells)
        for rule in columns:
            rule_key = rule.get("rule_ref") or rule.get("property_name") or "?"
            entry = r["cells"].get(rule_key)
            unit = rule.get("unit", "") or ""
            status = entry.get("status") if entry else None
            if entry is None:
                actual_txt = "—"
            else:
                actual_v = entry.get("actual")
                actual_txt = _fmt_val_s(actual_v) + (
                    f" {unit}" if unit and actual_v is not None else ""
                )
            cell_cls = (
                "bg-red-100 text-red-800 font-semibold"
                if status == "FAIL"
                else "bg-yellow-100 text-yellow-800 font-semibold"
                if status == "MISSING"
                else ""
            )
            tds.append(
                Td(
                    actual_txt,
                    cls=f"px-3 py-2 text-xs font-mono border-l-2 border-slate-300 {cell_cls}",
                )
            )
            tds.append(Td(_rule_required_text(rule), cls="px-3 py-2 text-xs text-muted-foreground"))
        issues_txt = "; ".join(r["issues"]) if r["issues"] else "✓ none"
        view_link = (
            A(
                "View in 3D",
                href=(
                    f"/viewer?project_id={_last_summary_project_id}"
                    f"&bcf_artifact_id={_bcf_artifact['id']}&element_guid={quote(r['guid'])}"
                ),
                target="_blank",
                cls="ml-2 text-blue-700 hover:underline whitespace-nowrap",
            )
            if r["issues"] and _bcf_artifact and r.get("guid")
            else ""
        )
        tds.append(
            Td(
                issues_txt,
                view_link,
                cls=f"px-3 py-2 text-xs {'text-red-700 font-medium' if r['issues'] else 'text-green-700'}",
            )
        )
        body_rows.append(Tr(*tds, cls="border-b border-muted last:border-0"))

    note = (
        P(
            f"Showing 50 of {len(rows)} elements — download for the full list.",
            cls="text-xs text-muted-foreground italic mt-2",
        )
        if len(rows) > 50
        else ""
    )

    return Div(
        Div(
            Div(
                H3("Summary — all elements tested", cls="text-sm font-semibold"),
                P(
                    "Red = fails its rule · Yellow = property missing · rows with any issue are tinted and sorted first.",
                    cls="text-[11px] text-muted-foreground",
                ),
            ),
            download_btns,
            cls="flex items-start justify-between mb-2 gap-4",
        ),
        Div(
            Table(
                Thead(Tr(*row1), Tr(*row2)),
                Tbody(*body_rows),
                cls="w-full text-xs",
            ),
            cls="overflow-auto border rounded-md max-h-96",
        ),
        note,
        cls="mb-4",
    )


def _egress_direction_select(rule_id, current: str):
    """Dropdown letting the user specify which direction the code treats as
    'egress' for door-swing-direction rules — no geometric swing-vs-path
    computation exists yet, so this just records the interpretation setting
    on the rule for that check to use once it's built."""
    current = (current or "outside").strip().lower()
    dom_id = f"egress-dir-{rule_id}"
    return Div(
        Span("Egress = ", cls="text-[11px] text-muted-foreground"),
        Select(
            Option("Outside", value="outside", selected=(current == "outside")),
            Option("Inside", value="inside", selected=(current == "inside")),
            id=dom_id,
            name="direction",
            cls="text-xs py-0.5",
            hx_post=f"/api/rules/{rule_id}/egress-direction",
            hx_trigger="change",
            hx_target=f"#{dom_id}-wrap",
            hx_swap="outerHTML",
        ),
        id=f"{dom_id}-wrap",
        cls="flex items-center gap-1",
    )


def _simple_rule_section(rule: dict):
    """Collapsible sub-section for one Module 4 rule check."""
    status = rule.get("status", "NO_ELEMENTS")
    if status == "NO_ELEMENTS":
        return ""
    ref = rule.get("rule_ref", "") or ""
    desc = (rule.get("rule_desc", "") or "")[:65]
    pass_c = rule.get("pass_count", 0)
    fail_c = rule.get("fail_count", 0)
    miss_c = rule.get("missing_count", 0)
    total = rule.get("total_count", 0)
    all_els = rule.get("all_elements", [])

    req_str = _rule_required_text(rule)

    summary_txt = (
        f"{fail_c} fail · {pass_c} pass · {miss_c} missing"
        if fail_c or miss_c
        else f"{pass_c}/{total} pass"
    )
    label = f"{ref}  {desc}" if ref else desc

    extra = None
    if (
        rule.get("rule_id")
        and (rule.get("property_name") or "").strip().lower() == "openingdirection"
    ):
        extra = _egress_direction_select(rule["rule_id"], rule.get("egress_direction"))

    return _collapsible_section(
        label,
        _simple_full_elem_tbl(all_els, rule),
        badge=f"{summary_txt}  ·  {req_str}",
        open=status in ("FAIL", "MISSING_DATA"),
        extra=extra,
    )


def _simple_domain_card(
    title: str,
    code_ref: str,
    rule_results: list,
    extra_sections: list | None = None,
    override_badge: tuple | None = None,
    category: str | None = None,
):
    """Outer collapsible card for one building code domain."""
    status_label, badge_cls = _domain_badge(rule_results)
    if override_badge:
        status_label, badge_cls = override_badge

    rule_secs = [_simple_rule_section(r) for r in rule_results]
    summary_tbl = _simple_summary_table(category, rule_results) if category else ""
    all_content = (
        ([summary_tbl] if summary_tbl else [])
        + [s for s in rule_secs if s]
        + (extra_sections or [])
    )

    has_fail = any(r.get("status") == "FAIL" for r in rule_results)
    if override_badge and "fail" in (override_badge[0] or "").lower():
        has_fail = True

    if not all_content:
        all_content = [
            P(
                "No applicable checks found in the rule library for this category.",
                cls="text-sm text-muted-foreground italic",
            )
        ]

    return _collapsible_card(
        title,
        Div(
            Span(
                status_label,
                cls=f"inline-block px-2 py-0.5 rounded text-xs font-semibold {badge_cls}",
            ),
            cls="flex items-center mb-4",
        ),
        P(code_ref, cls="text-xs text-muted-foreground -mt-2 mb-2") if code_ref else "",
        *all_content,
        open=has_fail,
    )


# ── Domain card builders ───────────────────────────────────────────────────────


def _simple_windows_card(by_class: dict, spatial_checks: dict):
    rules = by_class.get("IfcWindow", [])
    daylight = (spatial_checks or {}).get("daylight", [])
    extra = []
    override = None
    code_ctx = _code_rule_context()

    if daylight:
        daylight_ref = next(
            (_result_ref(r) for r in daylight if _result_ref(r)),
            code_ctx["daylight_ref"],
        )
        d_pass = sum(1 for r in daylight if r["passes"])
        d_fail = len(daylight) - d_pass
        d_rows = []
        for r in sorted(daylight, key=lambda x: x["passes"]):
            sc = "text-green-700 font-semibold" if r["passes"] else "text-red-700 font-semibold"
            d_rows.append(
                Tr(
                    Td(r.get("storey_name") or "—", cls="px-3 py-2 text-xs text-muted-foreground"),
                    Td(r["space_name"][:35], cls="px-3 py-2 text-xs"),
                    Td(f"{r['floor_area_m2']:.1f}", cls="px-3 py-2 text-xs font-mono text-right"),
                    Td(
                        f"{r['total_window_area_m2']:.2f}",
                        cls="px-3 py-2 text-xs font-mono text-right",
                    ),
                    Td(f"{r['daylight_ratio']:.3f}", cls="px-3 py-2 text-xs font-mono text-right"),
                    Td("✓ Pass" if r["passes"] else "✗ Fail", cls=f"px-3 py-2 text-xs {sc}"),
                    cls="border-b border-muted last:border-0",
                )
            )
        extra.append(
            _collapsible_section(
                _section_title(
                    "Daylight Ratio",
                    f"{daylight_ref} (>= {code_ctx['daylight_ratio_label']} floor area)",
                ),
                Div(
                    Table(
                        Thead(
                            Tr(
                                *[
                                    Th(
                                        h,
                                        cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                                    )
                                    for h in (
                                        "Floor",
                                        "Room",
                                        "Floor m²",
                                        "Window m²",
                                        "Ratio",
                                        "Status",
                                    )
                                ]
                            )
                        ),
                        Tbody(*d_rows),
                        cls="w-full text-xs",
                    ),
                    cls="overflow-auto border rounded-md max-h-64",
                ),
                badge=f"{d_pass}/{len(daylight)} pass",
                open=d_fail > 0,
            )
        )
        if d_fail > 0 and not any(r.get("status") == "FAIL" for r in rules):
            override = (f"{d_fail} room(s) below daylight ratio", "bg-red-100 text-red-800")

    return _simple_domain_card(
        "Windows & Glazing",
        code_ctx["daylight_ref"],
        rules,
        extra,
        override,
        category="windows",
    )


def _simple_doors_card(by_class: dict):
    code_ctx = _code_rule_context()
    return _simple_domain_card(
        "Doors",
        _domain_ref_for_targets(["IfcDoor"], code_ctx["exit_ref"]),
        by_class.get("IfcDoor", []),
        category="doors",
    )


def _simple_stairs_card(by_class: dict):
    rules = by_class.get("IfcStairFlight", []) + by_class.get("IfcRailing", [])
    code_ctx = _code_rule_context()
    return _simple_domain_card(
        "Stairs, Guards & Handrails",
        _domain_ref_for_targets(["IfcStairFlight", "IfcRailing"], code_ctx["daylight_ref"]),
        rules,
        category="stairs",
    )


def _simple_ramps_card(by_class: dict):
    rules = by_class.get("IfcRamp", []) + by_class.get("IfcRampFlight", [])
    code_ctx = _code_rule_context()
    return _simple_domain_card(
        "Ramps",
        _domain_ref_for_targets(["IfcRamp", "IfcRampFlight"], code_ctx["daylight_ref"]),
        rules,
        category="ramps",
    )


def _simple_egress_card(egress: dict):
    """Egress domain card built from Tier 3 egress check results."""
    if not egress:
        code_ctx = _code_rule_context()
        return _simple_domain_card("Means of Egress", code_ctx["travel_ref"], [])

    exit_data = egress.get("exit_count", {})
    travel = egress.get("travel_distance", [])
    total_exits = exit_data.get("total_exterior_doors", 0)
    exit_results = exit_data.get("results", [])
    code_ctx = _code_rule_context()
    exit_ref = next(
        (_result_ref(r) for r in exit_results if _result_ref(r)),
        code_ctx["exit_ref"],
    )
    travel_ref = next(
        (_result_ref(r) for r in travel if _result_ref(r)),
        code_ctx["travel_ref"],
    )

    all_passes = [r["passes"] for r in exit_results] + [r["passes"] for r in travel]
    if not all_passes:
        override = ("N/A", "bg-gray-100 text-gray-500")
    elif all(all_passes):
        override = ("All pass", "bg-green-100 text-green-800")
    else:
        fail_n = sum(1 for p in all_passes if not p)
        override = (f"{fail_n} check(s) failed", "bg-red-100 text-red-800")

    extra = []
    if exit_results:
        e_rows = []
        for r in exit_results:
            sc = "text-green-700 font-semibold" if r["passes"] else "text-red-700 font-semibold"
            e_rows.append(
                Tr(
                    Td(r["storey"], cls="px-3 py-2 text-xs"),
                    Td(str(r["exit_count"]), cls="px-3 py-2 text-xs font-mono text-center"),
                    Td(str(r["required_min"]), cls="px-3 py-2 text-xs font-mono text-center"),
                    Td("✓ Pass" if r["passes"] else "✗ Fail", cls=f"px-3 py-2 text-xs {sc}"),
                    cls="border-b border-muted last:border-0",
                )
            )
        extra.append(
            _collapsible_section(
                f"{_section_title('Exit Count', exit_ref)}  ({total_exits} exterior door(s))",
                Div(
                    Table(
                        Thead(
                            Tr(
                                *[
                                    Th(
                                        h,
                                        cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                                    )
                                    for h in ("Storey", "Exits", "Required", "Status")
                                ]
                            )
                        ),
                        Tbody(*e_rows),
                        cls="w-full text-xs",
                    ),
                    cls="overflow-auto border rounded-md",
                ),
                badge=f"{sum(1 for r in exit_results if r['passes'])}/{len(exit_results)} pass",
                open=any(not r["passes"] for r in exit_results),
            )
        )
    else:
        extra.append(
            _collapsible_section(
                _section_title("Exit Count", exit_ref),
                P(
                    "No exterior doors found. Tag doors as IsExternal=True in your authoring tool.",
                    cls="text-xs text-yellow-700",
                ),
                open=True,
            )
        )

    if travel:
        td_pass = sum(1 for r in travel if r["passes"])
        td_rows = []
        for r in sorted(travel, key=lambda x: (x["passes"], -(x.get("travel_distance_m") or 0))):
            dist = (
                f"{r['travel_distance_m']:.1f} m"
                if r.get("travel_distance_m") is not None
                else "No path"
            )
            sc = "text-green-700 font-semibold" if r["passes"] else "text-red-700 font-semibold"
            td_rows.append(
                Tr(
                    Td(r.get("storey_name") or "—", cls="px-3 py-2 text-xs text-muted-foreground"),
                    Td(r["space_name"][:35], cls="px-3 py-2 text-xs"),
                    Td(dist, cls="px-3 py-2 text-xs font-mono text-right"),
                    Td(r.get("nearest_exit") or "—", cls="px-3 py-2 text-xs"),
                    Td(
                        "✓ Pass"
                        if r["passes"]
                        else ("✗ No path" if r.get("no_path") else "✗ Exceeds"),
                        cls=f"px-3 py-2 text-xs {sc}",
                    ),
                    cls="border-b border-muted last:border-0",
                )
            )
        extra.append(
            _collapsible_section(
                _section_title("Travel Distance", travel_ref),
                Div(
                    Table(
                        Thead(
                            Tr(
                                *[
                                    Th(
                                        h,
                                        cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                                    )
                                    for h in ("Floor", "Room", "Distance", "Nearest Exit", "Status")
                                ]
                            )
                        ),
                        Tbody(*td_rows),
                        cls="w-full text-xs",
                    ),
                    cls="overflow-auto border rounded-md max-h-64",
                ),
                badge=f"{td_pass}/{len(travel)} pass",
                open=td_pass < len(travel),
            )
        )

    return _simple_domain_card("Means of Egress", travel_ref, [], extra, override)


def _simple_washrooms_card(by_class: dict, building_summary: dict):
    rules = by_class.get("IfcSanitaryTerminal", [])
    code_ctx = _code_rule_context()
    return _simple_domain_card(
        "Washrooms & Accessibility",
        _domain_ref_for_targets(["IfcSanitaryTerminal"], code_ctx["daylight_ref"]),
        rules,
        category="washrooms",
    )


def _simple_plumbing_card(building_summary: dict):
    fixture_counts = (building_summary or {}).get("fixture_counts", {})
    code_ctx = _code_rule_context()
    if not fixture_counts:
        return _simple_domain_card(
            "Plumbing Fixture Counts",
            _domain_ref_for_targets(["IfcSanitaryTerminal"], code_ctx["daylight_ref"]),
            [],
            override_badge=("N/A — no fixtures found", "bg-gray-100 text-gray-500"),
        )
    badges = [
        Span(
            f"{k}: {v}",
            cls="inline-block px-2 py-1 rounded-full text-xs bg-cyan-50 text-cyan-800 border border-cyan-200 font-medium",
        )
        for k, v in sorted(fixture_counts.items())
    ]
    return _simple_domain_card(
        "Plumbing Fixture Counts",
        _domain_ref_for_targets(["IfcSanitaryTerminal"], code_ctx["daylight_ref"]),
        [],
        extra_sections=[
            _collapsible_section(
                "Fixture Inventory",
                Div(*badges, cls="flex flex-wrap gap-2"),
                badge=f"{sum(fixture_counts.values())} fixtures",
                open=True,
            )
        ],
        override_badge=("Inventory only", "bg-blue-100 text-blue-800"),
    )


def _simple_fire_card(by_class: dict, spatial_checks: dict, building_summary: dict):
    rules = by_class.get("IfcAlarm", [])
    spatial = spatial_checks or {}
    fire_sep = spatial.get("fire_separation", [])
    alarm_counts = (building_summary or {}).get("alarm_counts", {})
    extra = []
    override = None
    code_ctx = _code_rule_context()

    if fire_sep:
        fire_ref = next(
            (_result_ref(r) for r in fire_sep if _result_ref(r)),
            code_ctx["fire_ref"],
        )
        f_pass = sum(1 for r in fire_sep if r["passes"])
        f_fail = len(fire_sep) - f_pass
        f_rows = []
        for r in sorted(fire_sep, key=lambda x: x["passes"]):
            rating_txt = r["fire_rating_raw"] or "⚠ Not declared"
            rc = (
                "text-red-700"
                if r["missing_rating"]
                else ("text-green-700" if r["passes"] else "text-orange-700")
            )
            spaces_txt = ", ".join(r["adjacent_spaces"][:2])
            if len(r["adjacent_spaces"]) > 2:
                spaces_txt += f" +{len(r['adjacent_spaces']) - 2}"
            f_rows.append(
                Tr(
                    Td(r["wall_name"][:35], cls="px-3 py-2 text-xs font-mono"),
                    Td(spaces_txt[:50], cls="px-3 py-2 text-xs"),
                    Td(rating_txt, cls=f"px-3 py-2 text-xs font-mono {rc}"),
                    Td("✓ Pass" if r["passes"] else "✗ Fail", cls=f"px-3 py-2 text-xs {rc}"),
                    cls="border-b border-muted last:border-0",
                )
            )
        extra.append(
            _collapsible_section(
                _section_title("Fire Separation", fire_ref),
                Div(
                    Table(
                        Thead(
                            Tr(
                                *[
                                    Th(
                                        h,
                                        cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                                    )
                                    for h in ("Wall", "Between Spaces", "Fire Rating", "Status")
                                ]
                            )
                        ),
                        Tbody(*f_rows),
                        cls="w-full text-xs",
                    ),
                    cls="overflow-auto border rounded-md max-h-64",
                ),
                badge=f"{f_pass}/{len(fire_sep)} pass",
                open=f_fail > 0,
            )
        )
        if f_fail > 0:
            override = (f"{f_fail} separation issue(s)", "bg-red-100 text-red-800")

    if alarm_counts:
        alarm_badges = [
            Span(
                f"{k}: {v}",
                cls="inline-block px-2 py-1 rounded-full text-xs bg-red-50 text-red-800 border border-red-200 font-medium",
            )
            for k, v in sorted(alarm_counts.items())
        ]
        extra.append(
            _collapsible_section(
                "Alarm Inventory",
                Div(*alarm_badges, cls="flex flex-wrap gap-2"),
                badge=f"{sum(alarm_counts.values())} alarms",
                open=False,
            )
        )

    return _simple_domain_card(
        "Fire Protection (House-Level)",
        _domain_ref_for_targets(["IfcAlarm", "IfcWall"], code_ctx["fire_ref"]),
        rules,
        extra,
        override,
        category="fire",
    )


def _simple_garage_card(spatial_checks: dict):
    spatial = spatial_checks or {}
    garage_sep = spatial.get("garage_separation", {})
    code_ctx = _code_rule_context()
    if not garage_sep:
        return _simple_domain_card(
            "Garage / Carport",
            code_ctx["garage_ref"],
            [],
            override_badge=("N/A — no garage detected", "bg-gray-100 text-gray-500"),
        )

    g_results = garage_sep.get("results", [])
    g_warnings = garage_sep.get("warnings", [])
    extra = []
    override = None

    if g_results:
        garage_ref = next(
            (_result_ref(r) for r in g_results if _result_ref(r)),
            code_ctx["garage_ref"],
        )
        g_pass = sum(1 for r in g_results if r["passes"])
        g_fail = len(g_results) - g_pass
        g_rows = []
        for r in sorted(g_results, key=lambda x: x["passes"]):
            rating_txt = r["fire_rating_raw"] or "⚠ Not declared"
            rc = (
                "text-red-700"
                if r["missing_rating"]
                else ("text-green-700" if r["passes"] else "text-orange-700")
            )
            g_rows.append(
                Tr(
                    Td(r["element_type"], cls="px-3 py-2 text-xs font-semibold"),
                    Td(r["element_name"][:35], cls="px-3 py-2 text-xs font-mono"),
                    Td(r["garage_space"][:25], cls="px-3 py-2 text-xs"),
                    Td(r["adjacent_space"][:25], cls="px-3 py-2 text-xs"),
                    Td(rating_txt, cls=f"px-3 py-2 text-xs font-mono {rc}"),
                    Td(f"≥ {r['required_min']} min", cls="px-3 py-2 text-xs font-mono"),
                    Td("✓ Pass" if r["passes"] else "✗ Fail", cls=f"px-3 py-2 text-xs {rc}"),
                    cls="border-b border-muted last:border-0",
                )
            )
        extra.append(
            _collapsible_section(
                _section_title("Garage Fire Separation", garage_ref),
                Div(
                    Table(
                        Thead(
                            Tr(
                                *[
                                    Th(
                                        h,
                                        cls="px-3 py-2 text-left text-xs font-semibold text-muted-foreground bg-muted",
                                    )
                                    for h in (
                                        "Type",
                                        "Element",
                                        "Garage Space",
                                        "Adjacent",
                                        "Rating",
                                        "Required",
                                        "Status",
                                    )
                                ]
                            )
                        ),
                        Tbody(*g_rows),
                        cls="w-full text-xs",
                    ),
                    cls="overflow-auto border rounded-md max-h-64",
                ),
                badge=f"{g_pass}/{len(g_results)} pass",
                open=g_fail > 0,
            )
        )
        if g_fail > 0:
            override = (f"{g_fail} separation issue(s)", "bg-red-100 text-red-800")
    elif g_warnings:
        garage_ref = code_ctx["garage_ref"]
        extra.append(
            _collapsible_section(
                _section_title("Garage Fire Separation", garage_ref),
                Div(*[Span(w, cls="block text-xs text-yellow-700") for w in g_warnings]),
                open=True,
            )
        )

    return _simple_domain_card("Garage / Carport", code_ctx["garage_ref"], [], extra, override)


def setup_routes(rt):
    """Register analysis workflow routes."""

    @rt("/analysis/ARCH")
    def analysis_simple():
        logger.info("ARCH page progress=0% step=page-requested path=/analysis/ARCH")
        projects = _projects_service.list_projects()
        logger.info("ARCH page progress=40%% step=projects-loaded count=%d", len(projects))
        documents = _documents_service.list_documents()
        logger.info("ARCH page progress=80%% step=documents-loaded count=%d", len(documents))
        logger.info("ARCH page progress=100% step=page-rendered path=/analysis/ARCH")
        return Title("ARCH Analysis - BIM Guard"), DashboardLayout(
            Container(_analysis_form(projects, documents, mode="simple"), cls="space-y-4")
        )

    @rt("/analysis/ARCH/results", methods=["POST"])
    async def analysis_simple_post(req: Request):
        logger.info("ARCH request progress=0% step=form-received path=/analysis/ARCH/results")
        form = await req.form()
        project_id_raw = form.get("project_id") or ""
        if not project_id_raw:
            logger.warning("ARCH request progress=0% step=validation-failed reason=missing-project")
            return Alert("Please select a project.", cls=AlertT.error)
        try:
            project_id = int(project_id_raw)
        except ValueError:
            logger.warning(
                "ARCH request progress=0%% step=validation-failed reason=invalid-project value=%r",
                project_id_raw,
            )
            return Alert("Invalid project selection.", cls=AlertT.error)

        rule_folder = (form.get("rule_folder") or "").strip()
        logger.info(
            "ARCH request progress=1%% step=validated project_id=%d rule_folder=%s",
            project_id,
            rule_folder or "all",
        )
        result = _bim_guard_app.orchestrate_workflow(
            project_id,
            doc_ids=[],
            analysis_theme="Architecture",
            rule_folder=rule_folder,
            include_openings=True,
            include_spaces=True,
            include_type_definitions=False,
        )
        if "error" in result:
            logger.warning(
                "ARCH request step=analysis-failed project_id=%d error=%s",
                project_id,
                result["error"],
            )
            return Alert(result["error"], cls=AlertT.error)

        from collections import defaultdict

        global _last_summary_project_id
        _last_summary_project_id = project_id

        # Persist a BCF artifact for this run — same mechanism _run_analysis_request
        # uses — so the "View in 3D" links below have a fresh, matching artifact to
        # point at (this route calls orchestrate_workflow directly rather than
        # through _run_analysis_request, so it doesn't get this for free).
        try:
            _report_artifact_service.persist_bcf(project_id, result.get("bcf_topics", []))
        except Exception:
            logger.exception("BCF persistence failed project_id=%d", project_id)

        logger.info("ARCH request progress=100%% step=results-rendering project_id=%d", project_id)
        project = result["project"]
        rule_compliance = result.get("rule_compliance", [])
        spatial_checks = result.get("spatial_checks", {})
        egress_checks = result.get("egress_checks", {})
        building_summary = result.get("building_summary", {})

        global _last_simple_compliance
        _last_simple_compliance = rule_compliance

        by_class: dict = defaultdict(list)
        for r in rule_compliance:
            by_class[r.get("target", "")].append(r)

        bldg_card = _building_summary_card(building_summary)
        result_folder = result.get("rule_folder", "")
        folder_note = f" (Folder: {result_folder})" if result_folder else ""

        sections = [
            Card(
                CardHeader(
                    CardTitle(f"{project.get('name', 'Project')} — ARCH Analysis{folder_note}")
                ),
                CardContent(
                    P(
                        f"Domain-based compliance check against {_active_code_refs_summary()}.",
                        cls="text-sm text-muted-foreground",
                    ),
                    Div(
                        _pass_rate_pill(result.get("rule_compliance_summary", {})),
                        cls="mt-3",
                    ),
                    _coverage_stat_line(result.get("rule_compliance_summary", {})),
                ),
            ),
            *(([bldg_card]) if bldg_card else []),
            _simple_windows_card(by_class, spatial_checks),
            _simple_doors_card(by_class),
            _simple_stairs_card(by_class),
            _simple_ramps_card(by_class),
            _simple_egress_card(egress_checks),
            _simple_washrooms_card(by_class, building_summary),
            _simple_plumbing_card(building_summary),
            _simple_fire_card(by_class, spatial_checks, building_summary),
            _simple_garage_card(spatial_checks),
        ]
        logger.info(
            "ARCH request progress=100%% step=response-ready project_id=%d sections=%d rule_results=%d",
            project_id,
            len(sections),
            len(rule_compliance),
        )
        return Div(*sections, cls="space-y-4")

    @rt("/api/rules/{rule_id}/egress-direction", methods=["POST"])
    async def api_set_egress_direction(rule_id: int, req: Request):
        """Save which direction ('outside'/'inside') a door-swing-direction
        rule should treat as egress. Interpretation setting only — no
        geometric swing-vs-path check consumes it yet."""
        form = await req.form()
        direction = (form.get("direction") or "outside").strip().lower()
        if direction not in ("outside", "inside"):
            direction = "outside"
        _rule_service.set_rule_parameter(rule_id, "egress_direction", direction)
        return _egress_direction_select(rule_id, direction)

    @rt("/analysis/initial")
    def analysis_initial():
        return RedirectResponse(url="/analysis/ARCH", status_code=303)

    @rt("/analysis/MEP")
    def analysis_run():
        projects = _projects_service.list_projects()
        documents = _documents_service.list_documents()

        return Title("MEP Analysis - BIM Guard"), DashboardLayout(
            Container(
                _mep_engine_rules_card(),
                _analysis_form(projects, documents, mode="model-rules"),
                cls="space-y-4",
            )
        )

    @rt("/analysis/initial/results", methods=["POST"])
    async def analysis_initial_post(req: Request):
        analysis_data, error_response = await _run_analysis_request(req)
        if error_response:
            return error_response

        project_id = analysis_data["project_id"]
        result = analysis_data["result"]
        project = result["project"]
        selected_theme = result.get("analysis_theme", "Architecture")
        rule_folder = result.get("rule_folder", "")
        theme_label = f"{selected_theme} Theme" + (
            f" · Folder: {rule_folder}" if rule_folder else ""
        )

        _last_initial_rule_compliance[project_id] = result.get("rule_compliance", [])

        bldg_card = _building_summary_card(result.get("building_summary", {}))
        spatial_card = _spatial_checks_card(result.get("spatial_checks", {}))
        egress_card = _egress_checks_card(result.get("egress_checks", {}))

        sections = [
            Card(
                CardHeader(CardTitle(f"{project.get('name', 'Project')} — {theme_label}")),
                CardContent(_build_ifc_summary_content(result, project_id)),
            ),
            *(([bldg_card]) if bldg_card else []),
            *(([spatial_card]) if spatial_card else []),
            *(([egress_card]) if egress_card else []),
            _build_ifc_graph_card(result, project_id),
        ]

        return Div(*sections, cls="space-y-4")

    @rt("/analysis/results", methods=["POST"])
    async def analysis_run_post(req: Request):
        analysis_data, error_response = await _run_analysis_request(req, forced_theme="MEP")
        if error_response:
            return error_response

        result = analysis_data["result"]
        project = result["project"]
        selected_theme = result.get("analysis_theme", "MEP")
        rule_folder = result.get("rule_folder", "")
        theme_label = f"{selected_theme} Theme" + (
            f" · Folder: {rule_folder}" if rule_folder else ""
        )

        compliance_card = _compliance_card(
            results=result.get("compliance_results", []),
            cost_impact=result.get("cost_impact"),
            issue_stats=result.get("issue_stats", {}),
            is_demo=result.get("compliance_is_demo", False),
            project_id=result.get("bcf_project_id"),
            error=result.get("compliance_error"),
        )

        rule_validation_card = _rule_validation_card(
            result.get("rule_validations", []), theme_label
        )

        rc = result.get("rule_compliance", [])
        rc_summary = result.get("rule_compliance_summary", {})
        rc_error = result.get("rule_compliance_error")

        # Cache for CSV download
        global _last_compliance_results
        _last_compliance_results = rc

        rule_compliance_card = _rule_compliance_card(
            rc, rc_summary, rc_error, theme_label, project_id=analysis_data["project_id"]
        )

        sections = [
            Card(
                CardHeader(CardTitle(f"{project.get('name', 'Project')} — {theme_label}")),
                CardContent(
                    P(
                        "Model loaded and ready for MEP corrosion analysis against the "
                        "saved MEP rule library.",
                        cls="text-sm text-muted-foreground",
                    )
                ),
            ),
            rule_validation_card,
        ]
        if rule_compliance_card:
            sections.append(rule_compliance_card)
        if selected_theme == "MEP" and compliance_card:
            sections.append(compliance_card)

        return Div(*sections, cls="space-y-4")

    @rt("/reports/compliance-csv")
    def compliance_csv_download():
        """Download the last rule compliance check as a CSV file."""
        from app.modules.module5_reporter import Module5_Reporter

        return StreamingResponse(
            Module5_Reporter().iter_csv_summary(_last_compliance_results),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="compliance_results.csv"'},
        )

    @rt("/reports/compliance-bcf")
    def compliance_bcf_download():
        """Download the last rule compliance check as a BCF 2.1 ZIP.

        One topic per failing element, GUID-linked — opens directly in
        Revit/Solibri/Navisworks/BIMcollab so the model author can jump
        straight to each flagged element instead of reading a table.
        """
        from starlette.responses import Response as StarletteResponse

        from app.modules.module5_reporter import Module5_Reporter

        bcf_bytes = Module5_Reporter().generate_bcf_zip(_last_compliance_results)
        return StarletteResponse(
            content=bcf_bytes,
            media_type="application/octet-stream",
            headers={"Content-Disposition": 'attachment; filename="compliance_results.bcf"'},
        )

    @rt("/reports/simple-summary-xlsx/{category}")
    def simple_summary_xlsx_download(category: str):
        """Download the full per-element summary table for one Simple Analysis category,
        with failing/missing cells highlighted the same way as the on-screen table."""
        import io

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from starlette.responses import Response as StarletteResponse

        targets = _SIMPLE_CATEGORY_TARGETS.get(category, [])
        rule_results = [r for r in _last_simple_compliance if r.get("target") in targets]
        columns, rows = _pivot_category_elements(rule_results)

        # Excel's standard "Bad" (red) / "Neutral" (yellow) conditional-format colours.
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fail_font = Font(color="9C0006")
        missing_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        missing_font = Font(color="9C6500")

        wb = Workbook()
        ws = wb.active
        ws.title = (category or "Summary")[:31]

        header = ["Element", "Floor", "Room", "GUID"]
        for rule in columns:
            label = f"{rule.get('property_name') or rule.get('rule_ref') or '?'} ({rule.get('rule_ref', '')})"
            header += [f"{label} — Actual", f"{label} — Required"]
        header.append("Issues")
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "E2"

        for r in rows:
            row_has_issue = bool(r["issues"])
            row = [r["name"], r["storey"], r["space"], r["guid"]]
            statuses = []
            for rule in columns:
                rule_key = rule.get("rule_ref") or rule.get("property_name") or "?"
                entry = r["cells"].get(rule_key)
                unit = rule.get("unit", "") or ""
                if entry is None:
                    actual_txt = "—"
                else:
                    actual_v = entry.get("actual")
                    actual_txt = _fmt_val_s(actual_v) + (
                        f" {unit}" if unit and actual_v is not None else ""
                    )
                row += [actual_txt, _rule_required_text(rule)]
                statuses.append(entry.get("status") if entry else None)
            row.append("; ".join(r["issues"]) if r["issues"] else "")
            ws.append(row)

            excel_row = ws.max_row
            if row_has_issue:
                ws.cell(row=excel_row, column=1).fill = fail_fill
                ws.cell(row=excel_row, column=len(row)).fill = fail_fill
            for i, status in enumerate(statuses):
                actual_col = 5 + (i * 2)  # A-D identity cols, then Actual/Required pairs
                if status == "FAIL":
                    cell = ws.cell(row=excel_row, column=actual_col)
                    cell.fill, cell.font = fail_fill, fail_font
                elif status == "MISSING":
                    cell = ws.cell(row=excel_row, column=actual_col)
                    cell.fill, cell.font = missing_fill, missing_font

        for col_cells in ws.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

        buf = io.BytesIO()
        wb.save(buf)
        return StarletteResponse(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{category or "summary"}_summary.xlsx"'
            },
        )

    @rt("/reports/simple-summary-pdf/{category}")
    def simple_summary_pdf_download(category: str):
        """Download a print-friendly PDF of the per-element summary.

        The wide actual/required pivot used on-screen and in Excel doesn't fit a
        printed page once a category has more than a handful of rules, so the PDF
        uses one compact Property/Actual/Required/Status table per element instead,
        with failing rows first and colour-highlighted the same way as the app.
        """
        import io

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            KeepTogether,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        from starlette.responses import Response as StarletteResponse

        targets = _SIMPLE_CATEGORY_TARGETS.get(category, [])
        rule_results = [r for r in _last_simple_compliance if r.get("target") in targets]
        columns, rows = _pivot_category_elements(rule_results)

        styles = getSampleStyleSheet()
        order = {"FAIL": 0, "MISSING": 1, "PASS": 2}

        story = [
            Paragraph(
                f"{(category or 'Summary').title()} — Compliance Summary", styles["Heading1"]
            ),
            Paragraph(
                f"{len(rows)} element(s) tested against {len(columns)} rule(s).", styles["Normal"]
            ),
            Spacer(1, 10),
        ]

        for r in rows:
            header_txt = f"{r['name']} — {r['storey']}"
            if r["space"] not in ("", "—"):
                header_txt += f" / {r['space']}"
            header_txt += f"   ({r['guid']})"

            props = []
            for rule in columns:
                rule_key = rule.get("rule_ref") or rule.get("property_name") or "?"
                entry = r["cells"].get(rule_key)
                status = entry.get("status") if entry else "MISSING"
                unit = rule.get("unit", "") or ""
                actual_txt = (
                    "—"
                    if entry is None
                    else (
                        _fmt_val_s(entry.get("actual"))
                        + (f" {unit}" if unit and entry.get("actual") is not None else "")
                    )
                )
                label = rule.get("property_name") or rule_key
                status_txt = {"PASS": "Pass", "FAIL": "Fail", "MISSING": "Missing"}.get(
                    status, status
                )
                props.append((label, actual_txt, _rule_required_text(rule), status_txt, status))
            props.sort(key=lambda p: order.get(p[4], 3))

            table_data = [["Property", "Actual", "Required", "Status"]] + [p[:4] for p in props]
            tbl = Table(table_data, colWidths=[160, 100, 100, 60])
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
            for i, (_, _, _, _, status) in enumerate(props, start=1):
                if status == "FAIL":
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFC7CE")))
                    style_cmds.append(("TEXTCOLOR", (0, i), (-1, i), colors.HexColor("#9C0006")))
                elif status == "MISSING":
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FFEB9C")))
                    style_cmds.append(("TEXTCOLOR", (0, i), (-1, i), colors.HexColor("#9C6500")))
            tbl.setStyle(TableStyle(style_cmds))

            story.append(
                KeepTogether(
                    [
                        Paragraph(header_txt, styles["Heading3"]),
                        tbl,
                        Spacer(1, 12),
                    ]
                )
            )

        if not rows:
            story.append(
                Paragraph("No elements found in the model for this category.", styles["Normal"])
            )

        buf = io.BytesIO()
        SimpleDocTemplate(
            buf,
            pagesize=letter,
            topMargin=36,
            bottomMargin=36,
            leftMargin=36,
            rightMargin=36,
        ).build(story)
        return StarletteResponse(
            content=buf.getvalue(),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{category or "summary"}_summary.pdf"'
            },
        )

    @rt("/reports/bcf/{project_id}")
    def bcf_download(project_id: int):
        artifact = _report_artifact_service.latest_bcf(project_id)
        if artifact is not None:
            bcf_path = _report_artifact_service.materialize(artifact)
            if bcf_path is not None and bcf_path.exists():
                filename = artifact.get("filename") or f"compliance_project_{project_id}.bcf"
                return FileResponse(
                    bcf_path,
                    media_type=artifact.get("content_type") or "application/octet-stream",
                    headers={
                        "Content-Disposition": f'attachment; filename="{filename}"'
                    },
                )

        bcf_file = os.path.join(_DATA_DIR, f"compliance_project_{project_id}.bcf")
        if not os.path.exists(bcf_file):
            return Alert(
                "BCF file not found. Run the analysis first to generate the report.",
                cls=AlertT.error,
            )
        return FileResponse(
            bcf_file,
            media_type="application/octet-stream",
            headers={
                "Content-Disposition": (
                    f'attachment; filename="compliance_project_{project_id}.bcf"'
                )
            },
        )

    @rt("/reports/bcf/artifacts/{artifact_id}")
    def bcf_artifact_download(artifact_id: int):
        """Download one persisted BCF export by artifact ID."""
        artifact = _report_artifact_service.get_bcf(artifact_id)
        if artifact is None:
            return Alert("BCF export not found.", cls=AlertT.error)

        bcf_path = _report_artifact_service.materialize(artifact)
        if bcf_path is None or not bcf_path.exists():
            return Alert("The stored BCF file could not be loaded.", cls=AlertT.error)

        filename = artifact.get("filename") or f"bcf_export_{artifact_id}.bcf"
        return FileResponse(
            bcf_path,
            media_type=artifact.get("content_type") or "application/octet-stream",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @rt("/reports/ifc-graph/{project_id}")
    def ifc_graph_report(project_id: int):
        """Standalone PyVis HTML page — embedded via iframe on the Initial Analysis results.

        Violation highlighting reuses the rule-compliance result already computed
        by the Initial Analysis page load (see _last_initial_rule_compliance) —
        re-running orchestrate_workflow() here measured 100s+ on a real model,
        all of it wasted just to get a list of failed-element GUIDs.
        """
        from app.modules.module2_ifc_read.ifc_graph import render_ifc_graph

        ifc_path = _projects_service.resolve_ifc_file(project_id)
        if not ifc_path:
            return Response(
                content="<p style='font-family:sans-serif;padding:2rem;'>"
                "No IFC file found for this project.</p>",
                media_type="text/html",
            )

        cached_compliance = _last_initial_rule_compliance.get(project_id, [])
        violations = [
            {"element": f.get("guid")}
            for r in cached_compliance
            for f in r.get("failures", [])
            if f.get("guid")
        ]

        try:
            graph_data = render_ifc_graph(ifc_path, violations)
        except Exception as exc:
            return Response(
                content=(
                    "<p style='font-family:sans-serif;padding:2rem;color:#b91c1c;'>"
                    f"Graph build failed: {exc}</p>"
                ),
                media_type="text/html",
            )

        return Response(content=graph_data["html"], media_type="text/html")

    @rt("/reports")
    def reports():
        artifacts = _report_artifact_service.list_bcf()
        projects_by_id = {
            int(project["id"]): project for project in _projects_service.list_projects()
        }

        rows = []
        for artifact in artifacts:
            artifact_id = int(artifact["id"])
            project_id = int(artifact["project_id"])
            project = projects_by_id.get(project_id, {})
            byte_size = int(artifact.get("byte_size") or 0)
            size_label = f"{byte_size / 1024:.1f} KB"
            created_at = str(artifact.get("created_at") or "-").replace("T", " ")
            rows.append(
                UiTableRow(
                    UiTableCell(project.get("name") or f"Project {project_id}"),
                    UiTableCell(artifact.get("filename") or f"BCF export {artifact_id}"),
                    UiTableCell(str(artifact.get("issue_count") or 0)),
                    UiTableCell(size_label),
                    UiTableCell(created_at),
                    UiTableCell(
                        Div(
                            IconLinkButton(
                                "scan-eye",
                                href=f"/viewer?project_id={project_id}&bcf_artifact_id={artifact_id}",
                                title="Visualize BCF in Viewer",
                            ),
                            IconLinkButton(
                                "download",
                                href=f"/reports/bcf/artifacts/{artifact_id}",
                                title="Download BCF",
                            ),
                            cls="flex items-center gap-1",
                        )
                    ),
                )
            )

        report_content = (
            UiTable(
                UiTableHeader(
                    UiTableRow(
                        UiTableHead("Project"),
                        UiTableHead("Export"),
                        UiTableHead("Issues"),
                        UiTableHead("Size"),
                        UiTableHead("Created"),
                        UiTableHead("Actions"),
                    )
                ),
                UiTableBody(*rows),
            )
            if rows
            else P(
                "No BCF exports are stored yet. Run an analysis to create one.",
                cls="text-sm text-muted-foreground",
            )
        )

        return Title("Reports - BIM Guard"), DashboardLayout(
            Container(
                Div(
                    H1("Reports", cls="text-3xl font-bold mb-4 tracking-tight"),
                    P(
                        "Review, download, and visualize every stored BCF compliance export.",
                        cls="text-muted-foreground mb-6",
                    ),
                    Card(
                        CardHeader(CardTitle(f"BCF exports ({len(artifacts)})")),
                        CardContent(report_content),
                    ),
                    cls="container mx-auto py-6",
                )
            )
        )
